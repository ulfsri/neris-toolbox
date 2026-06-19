"""
NERIS Entity Set Roster — Departments & Users
================================================
This script is for federal agencies to export all departments and users under their entity set. 
This script asks for your entity set NERIS ID, which is required for auth. 

Two sheets: 
  - USERS       : one row per user per department
  - DEPARTMENTS : one row per department with feature flags, onboarding
                  status, reporting status, and vendor/integration info
"""

import sys
import subprocess
import traceback
import time
from datetime import datetime

def ensure_dependencies():
    def pip_install(*packages):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *packages, "--quiet"],
            check=True
        )
    try:
        import openpyxl  # noqa
    except ImportError:
        print("Installing openpyxl...")
        pip_install("openpyxl")
        print("✓ openpyxl installed")

    try:
        from neris_api_client import NerisApiClient  # noqa
        print("✓ Dependencies ready")
    except ImportError:
        print("Installing neris-api-client...")
        pip_install("neris-api-client")
        print("✓ neris-api-client installed")

def prompt_config():
    print("\n" + "=" * 60)
    print("  NERIS Entity Set Roster — Departments & Users")
    print("=" * 60)

    print("\n── Credentials ──────────────────────────────────")
    entity_set_nuid = input("NERIS Entity Set ID: ").strip()
    username        = input("NERIS Email: ").strip()
    print("NERIS Password (note: characters will be visible):")
    password = input("> ").strip()

    if not entity_set_nuid or not username or not password:
        sys.exit("✗ Entity Set NUID, email, and password are all required.")

    return entity_set_nuid, username, password

BASE_URL = "https://api.neris.fsri.org/v1"


def authenticate(username, password):
    from neris_api_client import NerisApiClient, Config

    print("\nConnecting to NERIS API...")
    client = NerisApiClient(Config(
        base_url=BASE_URL,
        grant_type="password",
        username=username,
        password=password,
    ))
    print("\n" + "=" * 60)
    print("  CHECK YOUR EMAIL FOR THE MFA CODE")
    print("=" * 60)
    client.health()
    print("✓ Authentication successful!")
    return client

def _make_styles():
    from openpyxl.styles import Font, PatternFill, Border, Side
    header_fill = PatternFill(start_color="262F68", end_color="262F68", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin        = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, header_font, thin_border


def _hcell(ws, row, col, value, header_fill, header_font, thin_border):
    from openpyxl.styles import Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = header_fill
    c.font      = header_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thin_border
    return c


def _dcell(ws, row, col, value, thin_border):
    from openpyxl.styles import Font, Alignment
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin_border
    return c


def autofit(ws, headers, min_width=14, max_width=40):
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len    = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

def _parse_response(res):
    if not hasattr(res, "status_code"):
        return res
    if res.status_code == 403:
        raise PermissionError("403 Forbidden — session may have expired.")
    if res.status_code != 200:
        raise ValueError(f"HTTP {res.status_code}: {res.text[:200]}")
    if not res.text or not res.text.strip():
        raise ValueError(f"Empty response body (HTTP {res.status_code})")
    return res.json()


def _call_with_retry(fn, *args, retries=3, delay=3, **kwargs):
    for attempt in range(1, retries + 1):
        try:
            res = fn(*args, **kwargs)
            return _parse_response(res)
        except PermissionError:
            raise
        except Exception as e:
            if attempt < retries:
                print(f"\n  ⚠ Attempt {attempt} failed ({e}) — retrying in {delay}s...")
                time.sleep(delay)
            else:
                raise


def get_entity_set_departments(client, entity_set_nuid):
    """
    Hit GET /auth/entity_set/{nuid} directly to get the entity set's name,
    type, and member list — no per-sub attachment scoping involved.
    Returns (departments: list[dict with 'neris_id' key], entity_set_name: str)

    Only FD-prefixed member IDs are kept since the rest of the roster
    pipeline (get_entity, user_entity_membership, etc.) is department-scoped.
    """
    url = f"{BASE_URL}/auth/entity_set/{entity_set_nuid}"
    print(f"\nFetching entity set: {entity_set_nuid}")

    try:
        res  = _call_with_retry(client._session.get, url)
    except Exception as e:
        print(f"\n✗ Error fetching entity set: {e}")
        traceback.print_exc()
        return [], entity_set_nuid

    if not isinstance(res, dict):
        print("✗ Unexpected response format for entity set.")
        return [], entity_set_nuid

    name    = res.get("name", "")
    es_type = res.get("type", "")
    members = res.get("members", [])

    neris_ids = []
    seen      = set()
    for m in members:
        nid = m.get("neris_id", "")
        if nid and nid not in seen:
            seen.add(nid)
            neris_ids.append(nid)

    print(f"✓ Entity set: {name or entity_set_nuid} (type: {es_type or 'unknown'})")
    print(f"✓ {len(neris_ids)} member NERIS IDs found")

    dept_ids = [nid for nid in neris_ids if nid.startswith("FD")]
    skipped  = [nid for nid in neris_ids if not nid.startswith("FD")]
    if skipped:
        print(f"⚠ Skipping {len(skipped)} non-department entity ID(s): "
              f"{', '.join(skipped)}")

    departments = [{"neris_id": nid} for nid in dept_ids]
    print(f"✓ Total departments to process: {len(departments)}")
    return departments, (name or entity_set_nuid)


def get_vendor_names(client, neris_id):
    """
    Fetch integration_title values for a department via
    GET /account/enrollment/{neris_id} — paginated.
    Returns a comma-separated string of vendor/integration names.
    """
    session = client._session
    titles  = []
    page    = 1

    while True:
        try:
            res  = session.get(
                f"{BASE_URL}/account/enrollment/{neris_id}",
                params={"page_size": 100, "page_number": page}
            )
            data        = _parse_response(res)
            enrollments = data.get("enrollments", [])
            page_count  = data.get("page_count", 1)

            for e in enrollments:
                title = e.get("integration_title", "")
                if title and title not in titles:
                    titles.append(title)

            if page >= page_count:
                break
            page += 1

        except PermissionError:
            break
        except Exception as e:
            print(f"  ⚠ Could not fetch enrollments for {neris_id}: {e}")
            break

    return ", ".join(titles)


def get_users_for_entity(client, neris_id):
    """
    Fetch all users and their roles for a department via
    GET /entity/{neris_id}/user_entity_membership — paginated.
    Returns a list of row dicts ready for the USERS sheet.
    """
    session = client._session
    rows    = []
    page    = 1

    while True:
        try:
            res  = session.get(
                f"{BASE_URL}/entity/{neris_id}/user_entity_membership",
                params={"page_size": 100, "page_number": page}
            )
            data       = _parse_response(res)
            users      = data.get("users", [])
            page_count = data.get("page_count", 1)

            for u in users:
                roles      = u.get("roles", [])
                role_names = ", ".join(r.get("name", "") for r in roles if r.get("name"))
                status     = u.get("status", "")
                rows.append({
                    "sub":          u.get("sub", ""),
                    "given_name":   u.get("given_name", ""),
                    "family_name":  u.get("family_name", ""),
                    "email":        u.get("email", ""),
                    "active":       u.get("active"),
                    "logged_in":    "Yes" if status == "CONFIRMED" else "No",
                    "dept_neris_id": neris_id,
                    "dept_name":    "",
                    "role":         role_names,
                })

            if page >= page_count:
                break
            page += 1

        except Exception as e:
            print(f"  ⚠ Could not fetch users for {neris_id}: {e}")
            break

    return rows


def get_reporting_dept_ids(client, dept_ids, max_workers=20):
    """
    Check which departments have at least one incident using parallel
    single-record API calls via ThreadPoolExecutor.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    reporting = set()
    print(f"\nChecking reporting status for {len(dept_ids)} departments...")

    def check_one(neris_id):
        try:
            res = client.list_incidents(neris_id_entity=neris_id, page_size=1)
            if not isinstance(res, dict):
                res = res.json()
            return neris_id, len(res.get("incidents", [])) > 0
        except Exception:
            return neris_id, False

    completed = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(check_one, d): d for d in dept_ids}
        for future in as_completed(futures):
            neris_id, has = future.result()
            if has:
                reporting.add(neris_id)
            completed += 1
            if completed % 50 == 0:
                print(f"  {completed}/{len(dept_ids)} checked...")

    print(f"✓ {len(reporting)} departments have at least one incident")
    return reporting

USER_HEADERS = [
    "NERIS ID",
    "Department Name",
    "First Name",
    "Last Name",
    "Email",
    "Active?",
    "Logged in?",
    "Role",
]

DEPT_HEADERS_BASE = [
    "NERIS ID",
    "Department Name",
    "Department Type",
    "Onboarding Status",
    "Direct Reporting Active?",
    "No Activity Report Active?",
    "Reporting",
    "Vendor / Integration",
]


def write_users_sheet(ws, user_rows):
    header_fill, header_font, thin_border = _make_styles()
    ws.freeze_panes = "A2"
    for col, h in enumerate(USER_HEADERS, start=1):
        _hcell(ws, 1, col, h, header_fill, header_font, thin_border)
    for r, row in enumerate(user_rows, start=2):
        _dcell(ws, r, 1, row.get("dept_neris_id", ""), thin_border)
        _dcell(ws, r, 2, row.get("dept_name", ""), thin_border)
        _dcell(ws, r, 3, row.get("given_name", ""), thin_border)
        _dcell(ws, r, 4, row.get("family_name", ""), thin_border)
        _dcell(ws, r, 5, row.get("email", ""), thin_border)
        _dcell(ws, r, 6, "Yes" if row.get("active") else "No", thin_border)
        _dcell(ws, r, 7, row.get("logged_in", "No"), thin_border)
        _dcell(ws, r, 8, row.get("role", ""), thin_border)
    autofit(ws, USER_HEADERS)


def write_departments_sheet(ws, dept_rows, include_vendor):
    header_fill, header_font, thin_border = _make_styles()
    headers = [h for h in DEPT_HEADERS_BASE if h != "Vendor / Integration" or include_vendor]
    ws.freeze_panes = "A2"
    for col, h in enumerate(headers, start=1):
        _hcell(ws, 1, col, h, header_fill, header_font, thin_border)
    for r, row in enumerate(dept_rows, start=2):
        flags = row.get("feature_flags", {}) or {}
        _dcell(ws, r, 1, row.get("neris_id", ""), thin_border)
        _dcell(ws, r, 2, row.get("name", ""), thin_border)
        _dcell(ws, r, 3, row.get("department_type", ""), thin_border)
        _dcell(ws, r, 4, row.get("onboarding_status", ""), thin_border)
        _dcell(ws, r, 5, "Yes" if flags.get("allow_ui_incident_creation") else "No", thin_border)
        _dcell(ws, r, 6, "Yes" if flags.get("allow_ui_no_activity_report_creation") else "No", thin_border)
        _dcell(ws, r, 7, row.get("_has_incidents", ""), thin_border)
        if include_vendor:
            _dcell(ws, r, 8, row.get("_vendor_names", ""), thin_border)
    autofit(ws, headers)

def main():
    ensure_dependencies()

    entity_set_nuid, username, password = prompt_config()
    client = authenticate(username, password)

    departments, label = get_entity_set_departments(client, entity_set_nuid)
    if not departments:
        print("\n⚠ No departments found.")
        return

    print(f"\nFetching details for {len(departments)} departments...")
    dept_rows = []
    user_rows = []

    all_dept_ids       = [d.get("neris_id") or d.get("id", "") for d in departments]
    reporting_dept_ids = get_reporting_dept_ids(client, all_dept_ids)

    for i, dept in enumerate(departments, start=1):
        neris_id = dept.get("neris_id") or dept.get("id", "")
        print(f"  [{i}/{len(departments)}] {neris_id}")
        time.sleep(0.1)

        try:
            detail = client.get_entity(neris_id)
            if not isinstance(detail, dict):
                detail = detail.json()
        except Exception as e:
            print(f"  ⚠ Could not fetch entity detail for {neris_id}: {e}")
            detail = dept

        detail["_vendor_names"]  = get_vendor_names(client, neris_id)
        detail["_has_incidents"] = "Yes" if neris_id in reporting_dept_ids else "No"
        dept_rows.append(detail)

        dept_name = detail.get("name", "")
        rows = get_users_for_entity(client, neris_id)
        for row in rows:
            row["dept_name"] = dept_name
        user_rows.extend(rows)

    print(f"\n✓ {len(dept_rows)} departments | {len(user_rows)} users collected")

    include_vendor = any(d.get("_vendor_names", "") for d in dept_rows)

    from openpyxl import Workbook
    import pathlib

    wb       = Workbook()
    ws_users = wb.active
    ws_users.title = "USERS"
    write_users_sheet(ws_users, user_rows)

    ws_depts = wb.create_sheet("DEPARTMENTS")
    write_departments_sheet(ws_depts, dept_rows, include_vendor)

    date_str   = datetime.now().strftime("%Y-%m-%d")
    safe_label = "".join(c if c.isalnum() or c in " -_" else "" for c in label).strip()
    filename   = f"Departments and Users {safe_label} {date_str}.xlsx"
    docs_dir   = pathlib.Path.home() / "Documents"
    docs_dir.mkdir(exist_ok=True)
    filepath   = str(docs_dir / filename)

    wb.save(filepath)
    print(f"\n✓ Report saved: {filepath}")

    print("\n" + "=" * 60)
    print("  PROCESS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
