#!/usr/bin/env python3
"""
NERIS Monthly Incident Count Report — Entity Set
---------------------------------------------------
This script is for federal agencies to export a month-by-month count of incident reported by department within their entity set. 
This script asks for your entity set NERIS ID, which is required for auth. 
"""

import sys
import subprocess
from datetime import datetime
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed


def ensure_dependencies():
    def pip_install(*packages):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *packages, "--quiet"],
            check=True
        )

    try:
        import pandas   
        import openpyxl 
    except ImportError:
        print("Installing pandas and openpyxl...")
        pip_install("pandas", "openpyxl")
        print("✓ pandas and openpyxl installed")

    try:
        from neris_api_client import NerisApiClient 
    except ImportError:
        print("Installing neris-api-client...")
        pip_install("neris-api-client")
        print("✓ neris-api-client installed")


def prompt_config():
    print("\n" + "=" * 60)
    print("  NERIS Monthly Incident Count Report — Entity Set")
    print("=" * 60)
    print("\nNote: This report covers all data from 1 Jan 2025 onward.")
    print("      NAR = No-Activity Report (confirmed zero incidents).")
    print("      Blank = no data of either kind submitted.\n")

    print("-- Credentials --")
    entity_set_nuid = input("NNERIS Entity Set ID: ").strip()
    username        = input("NERIS Email: ").strip()
    print("NERIS Password (note: characters will be visible):")
    password = input("> ").strip()

    if not username or not password or not entity_set_nuid:
        sys.exit("✗ Entity Set NUID, email, and password are all required.")

    return entity_set_nuid, username, password



BASE_URL = "https://api.neris.fsri.org/v1"


def authenticate(username, password):
    from neris_api_client import NerisApiClient, Config

    print("\nConnecting to NERIS API...")
    client = NerisApiClient(Config(
        base_url=BASE_URL,
        grant_type="password",
        username=username,
        password=password,
    ))

    print("\n" + "=" * 60)
    print("  CHECK YOUR EMAIL FOR THE MFA CODE")
    print("=" * 60)

    # Trigger MFA
    client.health()

    print("\n✓ Authentication successful!")
    return client

def get_entity_set_entities(client, entity_set_nuid):
    url = f"{BASE_URL}/auth/entity_set/{entity_set_nuid}"
    print(f"\nFetching entity set: {entity_set_nuid}")

    try:
        res  = client._session.get(url)
        res.raise_for_status()
        data = res.json()
    except Exception as e:
        sys.exit(f"✗ Could not fetch entity set: {e}")

    if not isinstance(data, dict):
        sys.exit("✗ Unexpected response format for entity set.")

    name    = data.get("name", "")
    es_type = data.get("type", "")
    members = data.get("members", [])

    neris_ids = []
    seen      = set()
    for m in members:
        nid = m.get("neris_id", "")
        if nid and nid not in seen:
            seen.add(nid)
            neris_ids.append(nid)

    print(f"✓ Entity set: {name or entity_set_nuid} (type: {es_type or 'unknown'})")
    print(f"✓ {len(neris_ids)} member NERIS IDs found")

    dept_ids = [nid for nid in neris_ids if nid.startswith("FD")]
    skipped  = [nid for nid in neris_ids if not nid.startswith("FD")]
    if skipped:
        print(f"⚠ Skipping {len(skipped)} non-department entity ID(s): "
              f"{', '.join(skipped)}")

    if not dept_ids:
        sys.exit("✗ No fire department (FD) entities found in this entity set.")

    print(f"\nLooking up {len(dept_ids)} department names...")
    all_entities = {}
    for i, did in enumerate(dept_ids, 1):
        try:
            entity = client.get_entity(did)
            all_entities[did] = entity.get("name", "") if isinstance(entity, dict) else ""
        except Exception:
            all_entities[did] = ""
        if i % 25 == 0:
            print(f"  {i}/{len(dept_ids)} departments looked up...")
    print("✓ Department name lookup complete")

    return all_entities, (name or entity_set_nuid)


def _month_label(dt_or_str):
    """Convert a datetime or 'MM/YYYY' / ISO string to 'Mon-YYYY'. Returns None on failure."""
    if dt_or_str is None:
        return None
    if isinstance(dt_or_str, str):
        if "/" in dt_or_str and len(dt_or_str) <= 7:
            try:
                m, y = dt_or_str.split("/")
                return datetime(int(y), int(m), 1).strftime("%b-%Y")
            except Exception:
                pass
        try:
            dt_or_str = datetime.fromisoformat(dt_or_str.replace("Z", "+00:00"))
        except Exception:
            return None
    try:
        return dt_or_str.strftime("%b-%Y")
    except Exception:
        return None


def fetch_incidents_for_entity(client, neris_id_entity, page_size=100):
    """
    Fetch all incidents for a single department.
    Returns a dict: { 'Mon-YYYY': count }
    """
    month_counts = defaultdict(int)
    next_cursor  = None

    while True:
        kwargs = dict(neris_id_entity=neris_id_entity, page_size=page_size)
        if next_cursor:
            kwargs["cursor"] = next_cursor

        res = client.list_incidents(**kwargs)
        if not isinstance(res, dict):
            res = res.json()

        batch = res.get("incidents", [])
        if not batch:
            break

        for inc in batch:
            disp = inc.get("dispatch") or {}
            ts   = disp.get("call_create") or disp.get("call_create_start")
            lbl  = _month_label(ts)
            if lbl:
                month_counts[lbl] += 1

        next_cursor = res.get("next_cursor")
        if not next_cursor:
            break

    return dict(month_counts)


def fetch_nars_for_entity(client, neris_id_entity, page_size=100):
    """
    Fetch all no-activity reports for a single department.
    Returns a set of 'Mon-YYYY' labels.
    """
    nar_months  = set()
    next_cursor = None

    while True:
        params = dict(neris_id_entity=neris_id_entity, page_size=page_size)
        if next_cursor:
            params["cursor"] = next_cursor

        r   = client._session.get(f"{BASE_URL}/no_activity_report", params=params)
        res = r.json()

        batch = res.get("reports", [])
        if not batch:
            break

        for report in batch:
            lbl = _month_label(report.get("month_year", ""))
            if lbl:
                nar_months.add(lbl)

        next_cursor = res.get("next_cursor")
        if not next_cursor:
            break

    return nar_months


def generate_month_columns():
    """All month labels from Jan-2025 through the current month."""
    cols, d = [], datetime(2025, 1, 1)
    now = datetime.now()
    while d <= now:
        cols.append(d.strftime("%b-%Y"))
        m = d.month + 1
        d = datetime(d.year + (m // 13), ((m - 1) % 12) + 1, 1)
    return cols

def export_to_excel(df_pivot, month_cols, label, inc_counts, nar_flags):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pathlib

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(c if c.isalnum() or c in " -_" else "" for c in label).strip()
    filename   = f"neris_monthly_counts_{safe_label}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Incident Counts"

    header_cols = ["NERIS Entity ID", "Department Name"] + month_cols
    ws.append(header_cols)

    hdr_fill     = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    nar_fill     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_fill   = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    hdr_font     = Font(color="FFFFFF", bold=True, size=11)
    total_font   = Font(bold=True, size=11)
    hdr_aln      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr_aln      = Alignment(horizontal="center", vertical="center")
    left_aln     = Alignment(horizontal="left",   vertical="center")
    thin_side    = Side(style="thin", color="BFBFBF")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side,  bottom=thin_side)

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_aln

    total_row_idx = len(df_pivot) + 1 
    rows = df_pivot.values.tolist()
    for r_idx, row_data in enumerate(rows, start=2):
        is_total = (r_idx == total_row_idx)
        for c_idx, val in enumerate(row_data, start=1):
            cell        = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border

            if is_total:
                cell.fill      = total_fill
                cell.font      = total_font
                cell.alignment = ctr_aln
            elif c_idx <= 2:
                cell.alignment = left_aln
            elif val == "NAR":
                cell.fill      = nar_fill
                cell.font      = Font(italic=True, color="7F6000")
                cell.alignment = ctr_aln
            elif val == "":
                cell.fill      = no_data_fill
                cell.alignment = ctr_aln
            else:
                cell.alignment = ctr_aln

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    for i in range(3, len(header_cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 11

    ws.freeze_panes = "C2"

    ws_legend  = wb.create_sheet("Legend")
    legend_data = [
        ("Symbol",       "Meaning"),
        ("(number)",     "Count of incidents reported for that department in that month"),
        ("NAR",          "No-Activity Report filed — department confirmed zero incidents"),
        ("(grey blank)", "No incidents and no no-activity report submitted for that month"),
        ("TOTAL row",    "Sum of numeric incident counts across all departments per month"),
    ]
    for lr in legend_data:
        ws_legend.append(lr)
    for cell in ws_legend[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    ws_legend.column_dimensions["A"].width = 14
    ws_legend.column_dimensions["B"].width = 75

    docs_dir = pathlib.Path.home() / "Documents"
    docs_dir.mkdir(exist_ok=True)
    filepath = str(docs_dir / filename)
    wb.save(filepath)
    return filepath


def main():
    ensure_dependencies()

    import pandas as pd

    entity_set_nuid, username, password = prompt_config()

    client = authenticate(username, password)

    all_entities, label = get_entity_set_entities(client, entity_set_nuid)
    all_eids            = list(all_entities.keys())

    print(f"\nFetching incidents and NARs for {len(all_eids)} departments in parallel...")
    print("(progress updates every 50 departments)\n")

    inc_counts = {}   
    nar_flags  = {}   
    completed  = 0
    errors     = []

    _first_dept_done = {"done": False}

    def fetch_dept_data(eid):
        counts = fetch_incidents_for_entity(client, eid)
        nars   = fetch_nars_for_entity(client, eid)
        if not _first_dept_done["done"]:
            _first_dept_done["done"] = True
            print(f"\n  [Diagnostic] First dept {eid}: "
                  f"{sum(counts.values())} incidents, {len(nars)} NAR months")
        return eid, counts, nars

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_dept_data, eid): eid for eid in all_eids}

        first_errors_shown = 0
        for future in as_completed(futures):
            eid = futures[future]
            try:
                eid, counts, nars = future.result()
                inc_counts[eid]   = counts
                nar_flags[eid]    = nars
            except Exception as e:
                errors.append((eid, str(e)))
                inc_counts[eid] = {}
                nar_flags[eid]  = set()
                if first_errors_shown < 3:
                    print(f"\n  ⚠ ERROR for {eid}: {e}")
                    first_errors_shown += 1

            completed += 1
            if completed % 50 == 0 or completed == len(all_eids):
                total_incidents = sum(sum(v.values()) for v in inc_counts.values())
                total_nars      = sum(len(v) for v in nar_flags.values())
                print(f"  {completed}/{len(all_eids)} departments complete "
                      f"| {total_incidents} incidents | {total_nars} NAR months")

    if errors:
        print(f"\n⚠ {len(errors)} department(s) had fetch errors:")
        for eid, err in errors[:10]:
            print(f"    {eid}: {err}")
        if len(errors) > 10:
            print(f"    ... and {len(errors) - 10} more")

    total_incidents = sum(sum(v.values()) for v in inc_counts.values())
    total_nars      = sum(len(v) for v in nar_flags.values())
    print(f"\n{'='*50}")
    print(f"Total incidents retrieved : {total_incidents}")
    print(f"Total NAR months on record: {total_nars}")
    print(f"{'='*50}")

    month_cols = generate_month_columns()
    rows = []

    for eid in sorted(all_eids):
        row = {
            "NERIS Entity ID": eid,
            "Department Name": all_entities.get(eid, ""),
        }
        dept_counts = inc_counts.get(eid, {})
        dept_nars   = nar_flags.get(eid, set())

        for mc in month_cols:
            if mc in dept_counts:
                row[mc] = dept_counts[mc]
            elif mc in dept_nars:
                row[mc] = "NAR"
            else:
                row[mc] = ""
        rows.append(row)

    df_pivot = pd.DataFrame(rows)

    summary = {"NERIS Entity ID": "TOTAL", "Department Name": ""}
    for mc in month_cols:
        nums = pd.to_numeric(df_pivot[mc], errors="coerce").dropna()
        summary[mc] = int(nums.sum()) if not nums.empty else ""
    df_pivot = pd.concat([df_pivot, pd.DataFrame([summary])], ignore_index=True)

    print(f"\nPivot table: {len(df_pivot)-1} departments x {len(month_cols)} months")

    filename = export_to_excel(df_pivot, month_cols, label, inc_counts, nar_flags)

    print(f"\n✓ Exported: {filename}")
    print(f"  Departments : {len(df_pivot)-1}")
    print(f"  Months      : {len(month_cols)}  ({month_cols[0]} - {month_cols[-1]})")
    print(f"  Incidents   : {total_incidents}")
    print(f"  NAR months  : {total_nars}")
    print("\nKey:")
    print("  (number)     = incident count")
    print("  NAR          = no-activity report filed (confirmed 0 incidents)")
    print("  (grey blank) = no data submitted for that month")

    print("\n" + "=" * 60)
    print("  PROCESS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
