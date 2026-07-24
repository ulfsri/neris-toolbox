"""
NERIS State Roster — Departments & Users
=========================================
Pulls all departments and users for a given state and exports a formatted
Excel workbook with two sheets:
  - USERS       — one row per user per department
  - DEPARTMENTS — one row per department with feature flags, reporting status,
                  and vendor/integration info
"""

import sys
import subprocess
import os
import time
import traceback
from datetime import datetime


def ensure_dependencies():
    def pip_install(*packages):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *packages, "--quiet"],
            check=True
        )
    try:
        import openpyxl  
    except ImportError:
        print("Installing openpyxl...")
        pip_install("openpyxl")
        print("✓ openpyxl installed")

    try:
        from neris_api_client import NerisApiClient 
        print("✓ Dependencies ready")
    except ImportError:
        print("Installing neris-api-client...")
        pip_install(
            "https://github.com/ulfsri/neris-api-client/archive/refs/heads/main.zip"
        )
        print("✓ neris-api-client installed")

ensure_dependencies()

from neris_api_client import NerisApiClient, Config
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



print("\n" + "=" * 60)
print("  NERIS State Roster — Departments & Users")
print("=" * 60)

print("\n── Credentials ──────────────────────────────────")
username   = input("NERIS Email: ").strip()
print("NERIS Password (note: characters will be visible):")
password   = input("> ").strip()

print("\n── Query Parameters ─────────────────────────────")
state_code = input("State Code (e.g. VA, MI, CA): ").strip().upper()

print(f"\n✓ Username:   {username}")
print(f"✓ Password:   {'*' * len(password)}")
print(f"✓ State Code: {state_code}")

if not username or not password:
    sys.exit("✗ Email and password are required.")
if not state_code:
    sys.exit("✗ State code is required.")


print("\nConnecting to NERIS API...")
client = NerisApiClient(Config(
    base_url="https://api.neris.fsri.org/v1",
    grant_type="password",
    username=username,
    password=password,
))
print("✓ Authentication successful!")


HEADER_FILL = PatternFill(start_color="262F68", end_color="262F68", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN        = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hcell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    return c


def _dcell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = THIN_BORDER
    return c


def autofit(ws, headers, min_width=14, max_width=40):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _parse_response(res):
    """Parse a Response object, raising a clear error on non-200 or empty body."""
    if not hasattr(res, "status_code"):
        return res  # already a dict/list
    if res.status_code == 403:
        raise PermissionError("403 Forbidden — session may have expired. Re-run the script.")
    if res.status_code != 200:
        raise ValueError(f"HTTP {res.status_code}: {res.text[:200]}")
    if not res.text or not res.text.strip():
        raise ValueError(f"Empty response body (HTTP {res.status_code})")
    return res.json()


def _call_with_retry(fn, *args, retries=3, delay=3, **kwargs):
    """Call fn(*args, **kwargs) with retry on transient failures."""
    for attempt in range(1, retries + 1):
        try:
            res = fn(*args, **kwargs)
            return _parse_response(res)
        except PermissionError:
            raise
        except Exception as e:
            if attempt < retries:
                print(f"\n  ⚠ Attempt {attempt} failed ({e}) — retrying in {delay}s...")
                time.sleep(delay)
            else:
                raise


def get_entity_departments(client, state_code):
    """Page through all departments for the given state code."""
    departments = []
    page_number = 1
    page_count  = None

    print(f"\nFetching departments for state: {state_code}")
    while True:
        print(f"  Page {page_number}... ", end="", flush=True)
        try:
            res = _call_with_retry(
                client.list_entities,
                page_size=100,
                page_number=page_number,
                state=state_code
            )

            if page_number == 1:
                page_count  = res.get("page_count") or 1
                total_count = res.get("total_count") or 0
                print(f"\n  [info] total_count={total_count}, page_count={page_count}")

            entities = res.get("entities", [])
            if not entities:
                print("empty — done")
                break

            departments.extend(entities)
            print(f"{len(entities)} retrieved (total so far: {len(departments)})")

            if page_number >= page_count:
                print("  All pages retrieved.")
                break

            page_number += 1

        except Exception as e:
            print(f"\n✗ Error on page {page_number}: {e}")
            traceback.print_exc()
            break

    print(f"✓ Total departments fetched: {len(departments)}")
    return departments


def get_vendor_names(client, neris_id):
    """
    Fetch integration_title values for a department via
    GET /account/enrollment/{neris_id} — paginated.
    Returns a comma-separated string of vendor/integration names.
    """
    base_url = "https://api.neris.fsri.org/v1"
    session  = client._session
    titles   = []
    page     = 1

    while True:
        try:
            res  = session.get(
                f"{base_url}/account/enrollment/{neris_id}",
                params={"page_size": 100, "page_number": page}
            )
            data        = _parse_response(res)
            enrollments = data.get("enrollments", [])
            page_count  = data.get("page_count", 1)

            for e in enrollments:
                title = e.get("integration_title", "")
                if title and title not in titles:
                    titles.append(title)

            if page >= page_count:
                break
            page += 1

        except PermissionError:
            break  
        except Exception as e:
            print(f"  ⚠ Could not fetch enrollments for {neris_id}: {e}")
            break

    return ", ".join(titles)


def get_users_for_entity(client, neris_id):
    """
    Fetch all users and their roles for a department via
    GET /entity/{neris_id}/user_entity_membership — paginated.
    Returns a list of row dicts ready for the USERS sheet.
    """
    base_url = "https://api.neris.fsri.org/v1"
    session  = client._session
    rows     = []
    page     = 1

    while True:
        try:
            res  = session.get(
                f"{base_url}/entity/{neris_id}/user_entity_membership",
                params={"page_size": 100, "page_number": page}
            )
            data       = _parse_response(res)
            users      = data.get("users", [])
            page_count = data.get("page_count", 1)

            for u in users:
                roles      = u.get("roles", [])
                role_names = ", ".join(r.get("name", "") for r in roles if r.get("name"))
                status     = u.get("status", "")
                rows.append({
                    "sub":           u.get("sub", ""),
                    "given_name":    u.get("given_name", ""),
                    "family_name":   u.get("family_name", ""),
                    "email":         u.get("email", ""),
                    "active":        u.get("active"),
                    "logged_in":     "Yes" if status == "CONFIRMED" else "No",
                    "dept_neris_id": neris_id,
                    "dept_name":     "",  
                    "role":          role_names,
                })

            if page >= page_count:
                break
            page += 1

        except Exception as e:
            print(f"  ⚠ Could not fetch users for {neris_id}: {e}")
            break

    return rows


def get_reporting_dept_ids(client, dept_ids, max_workers=20):
    """
    Check which departments have at least one incident using parallel
    single-record API calls via ThreadPoolExecutor.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    reporting = set()
    print(f"\nChecking reporting status for {len(dept_ids)} departments...")

    def check_one(neris_id):
        try:
            res = client.list_incidents(neris_id_entity=neris_id, page_size=1)
            if not isinstance(res, dict):
                res = res.json()
            return neris_id, len(res.get("incidents", [])) > 0
        except Exception:
            return neris_id, False

    completed = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(check_one, d): d for d in dept_ids}
        for future in as_completed(futures):
            neris_id, has = future.result()
            if has:
                reporting.add(neris_id)
            completed += 1
            if completed % 50 == 0:
                print(f"  {completed}/{len(dept_ids)} checked...")

    print(f"✓ {len(reporting)} departments have at least one incident")
    return reporting


USER_HEADERS = [
    "NERIS ID",
    "Department Name",
    "First Name",
    "Last Name",
    "Email",
    "Active?",
    "Logged in?",
    "Role",
]

DEPT_HEADERS_BASE = [
    "NERIS ID",
    "Department Name",
    "Department Type",
    "Onboarding Status",
    "Direct Reporting Active?",
    "No Activity Report Active?",
    "Reporting",
    "Vendor / Integration",  # conditionally dropped if all blank
]


def write_users_sheet(ws, user_rows):
    ws.freeze_panes = "A2"
    for col, h in enumerate(USER_HEADERS, start=1):
        _hcell(ws, 1, col, h)
    for r, row in enumerate(user_rows, start=2):
        _dcell(ws, r, 1, row.get("dept_neris_id", ""))
        _dcell(ws, r, 2, row.get("dept_name", ""))
        _dcell(ws, r, 3, row.get("given_name", ""))
        _dcell(ws, r, 4, row.get("family_name", ""))
        _dcell(ws, r, 5, row.get("email", ""))
        _dcell(ws, r, 6, "Yes" if row.get("active") else "No")
        _dcell(ws, r, 7, row.get("logged_in", "No"))
        _dcell(ws, r, 8, row.get("role", ""))
    autofit(ws, USER_HEADERS)


def write_departments_sheet(ws, dept_rows, include_vendor):
    headers = [h for h in DEPT_HEADERS_BASE if h != "Vendor / Integration" or include_vendor]
    ws.freeze_panes = "A2"
    for col, h in enumerate(headers, start=1):
        _hcell(ws, 1, col, h)
    for r, row in enumerate(dept_rows, start=2):
        flags = row.get("feature_flags", {}) or {}
        _dcell(ws, r, 1, row.get("neris_id", ""))
        _dcell(ws, r, 2, row.get("name", ""))
        _dcell(ws, r, 3, row.get("department_type", ""))
        _dcell(ws, r, 4, row.get("onboarding_status", ""))
        _dcell(ws, r, 5, "Yes" if flags.get("allow_ui_incident_creation") else "No")
        _dcell(ws, r, 6, "Yes" if flags.get("allow_ui_no_activity_report_creation") else "No")
        _dcell(ws, r, 7, row.get("_has_incidents", ""))
        if include_vendor:
            _dcell(ws, r, 8, row.get("_vendor_names", ""))
    autofit(ws, headers)




departments = get_entity_departments(client, state_code)

if not departments:
    print("\n⚠ No departments found.")
else:
    print(f"\nFetching details for {len(departments)} departments...")
    dept_rows = []
    user_rows = []

    all_dept_ids = [
        d.get("neris_id") or d.get("id", "") for d in departments
    ]
    reporting_dept_ids = get_reporting_dept_ids(client, all_dept_ids)

    for i, dept in enumerate(departments, start=1):
        neris_id  = dept.get("neris_id") or dept.get("id", "")
        dept_name = dept.get("name", "")
        print(f"  [{i}/{len(departments)}] {dept_name} ({neris_id})")
        time.sleep(0.1)

        try:
            detail = client.get_entity(neris_id)
            if not isinstance(detail, dict):
                detail = detail.json()
        except Exception as e:
            print(f"  ⚠ Could not fetch entity detail for {neris_id}: {e}")
            detail = dept

        detail["_vendor_names"]  = get_vendor_names(client, neris_id)
        detail["_has_incidents"] = "Yes" if neris_id in reporting_dept_ids else "No"
        dept_rows.append(detail)

        rows = get_users_for_entity(client, neris_id)
        for row in rows:
            row["dept_name"] = dept_name
        user_rows.extend(rows)

    print(f"\n✓ {len(dept_rows)} departments | {len(user_rows)} users collected")

    include_vendor = any(d.get("_vendor_names", "") for d in dept_rows)

    wb       = Workbook()
    ws_users = wb.active
    ws_users.title = "USERS"
    write_users_sheet(ws_users, user_rows)

    ws_depts = wb.create_sheet("DEPARTMENTS")
    write_departments_sheet(ws_depts, dept_rows, include_vendor)

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"Departments and Users {state_code} {date_str}.xlsx"

    wb.save(filename)
    print(f"\n✓ Report saved: {filename}")

print("\n" + "=" * 60)
print("✓ PROCESS COMPLETE")
print("=" * 60)
