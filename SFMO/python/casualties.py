"""
NERIS Fire Casualties Report
====================================================
Returns one row per casualty (death or injury) for incidents where at least
one incident type falls under a selected FIRE domain.
"""

import sys
import subprocess
import os
import time
from getpass import getpass
from datetime import datetime, timezone

print("Installing NERIS API client...")
try:
    result = subprocess.run(
        [sys.executable, "-m", "pip", "install",
         "https://github.com/ulfsri/neris-api-client/archive/refs/heads/main.zip",
         "--quiet"],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        print("✓ NERIS API client installed successfully")
    else:
        print(f"Installation output: {result.stdout}")
        print(f"Installation errors: {result.stderr}")
except Exception as e:
    print(f"Installation error: {e}")

try:
    from neris_api_client import NerisApiClient
    print("✓ NERIS API Client loaded")
except ImportError:
    print("✗ NERIS API Client not found. Exiting.")
    sys.exit(1)

import pandas as pd
try:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

FIRE_DOMAINS = [
    "FIRE||STRUCTURE_FIRE",
    "FIRE||OUTSIDE_FIRE",
    "FIRE||SPECIAL_FIRE",
    "FIRE||TRANSPORTATION_FIRE",
]

FIRE_DOMAIN_LEAF_TYPES = {
    "FIRE||OUTSIDE_FIRE": [
        "FIRE||OUTSIDE_FIRE||CONSTRUCTION_WASTE",
        "FIRE||OUTSIDE_FIRE||DUMPSTER_OUTDOOR_CONTAINER_FIRE",
        "FIRE||OUTSIDE_FIRE||OTHER_OUTSIDE_FIRE",
        "FIRE||OUTSIDE_FIRE||OUTSIDE_TANK_FIRE",
        "FIRE||OUTSIDE_FIRE||TRASH_RUBBISH_FIRE",
        "FIRE||OUTSIDE_FIRE||UTILITY_INFRASTRUCTURE_FIRE",
        "FIRE||OUTSIDE_FIRE||VEGETATION_GRASS_FIRE",
        "FIRE||OUTSIDE_FIRE||WILDFIRE_URBAN_INTERFACE",
        "FIRE||OUTSIDE_FIRE||WILDFIRE_WILDLAND",
    ],
    "FIRE||SPECIAL_FIRE": [
        "FIRE||SPECIAL_FIRE||ESS_FIRE",
        "FIRE||SPECIAL_FIRE||EXPLOSION",
        "FIRE||SPECIAL_FIRE||INFRASTRUCTURE_FIRE",
    ],
    "FIRE||STRUCTURE_FIRE": [
        "FIRE||STRUCTURE_FIRE||CHIMNEY_FIRE",
        "FIRE||STRUCTURE_FIRE||CONFINED_COOKING_APPLIANCE_FIRE",
        "FIRE||STRUCTURE_FIRE||ROOM_AND_CONTENTS_FIRE",
        "FIRE||STRUCTURE_FIRE||STRUCTURAL_INVOLVEMENT_FIRE",
    ],
    "FIRE||TRANSPORTATION_FIRE": [
        "FIRE||TRANSPORTATION_FIRE||AIRCRAFT_FIRE",
        "FIRE||TRANSPORTATION_FIRE||BOAT_PERSONAL_WATERCRAFT_BARGE_FIRE",
        "FIRE||TRANSPORTATION_FIRE||POWERED_MOBILITY_DEVICE_FIRE",
        "FIRE||TRANSPORTATION_FIRE||TRAIN_RAIL_FIRE",
        "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_COMMERCIAL",
        "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_FOOD_TRUCK",
        "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_PASSENGER",
        "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_RV",
    ],
}

REQUEST_DELAY_SECONDS = 0.3
MAX_RETRIES = 6
RETRY_BACKOFF_SECONDS = 5
CHECKPOINT_EVERY_N_PAGES = 25
CHECKPOINT_PATH = "neris_fire_casualties_checkpoint.csv"


def _fetch_page(client, kwargs):
    """Call list_incidents, retrying with backoff on transient errors."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.list_incidents(**kwargs)
            if not isinstance(resp, dict):
                status = getattr(resp, "status_code", "?")
                body = getattr(resp, "text", "")[:200]
                raise RuntimeError(f"Non-JSON response (status {status}): {body}")
            return resp
        except Exception as e:
            if attempt == MAX_RETRIES:
                raise
            wait = RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1))
            print(f"  ⚠ Page fetch failed ({e}) — retrying in {wait}s "
                  f"(attempt {attempt}/{MAX_RETRIES})")
            time.sleep(wait)


def incident_domain(type_str: str) -> str:
    parts = type_str.split("||")
    return "||".join(parts[:2]) if len(parts) >= 2 else type_str


def incident_matches(incident: dict, selected_domains: set) -> bool:
    for it in incident.get("incident_types", []) or []:
        t = it.get("type", "")
        if incident_domain(t) in selected_domains:
            return True
    return False


def ordered_incident_types(incident: dict, n: int = 3):
    types = incident.get("incident_types", []) or []
    types_sorted = sorted(types, key=lambda x: not x.get("primary", False))
    values = [t.get("type", "") for t in types_sorted]
    values += [""] * (n - len(values))
    return values[:n]

print("\n" + "=" * 70)
print("  NERIS Fire Casualties Report — Login")
print("=" * 70)

NERIS_USERNAME = input("NERIS Email:      ").strip()
NERIS_PASSWORD = getpass("NERIS Password:   ")
STATE_CODE     = input("State Code (e.g. VA): ").strip().upper()
ENTITY_ID      = input("Entity NERIS ID (optional, press Enter to skip): ").strip()

print(f"\n✓ Username:   {NERIS_USERNAME}")
print(f"✓ Password:   {'*' * len(NERIS_PASSWORD)}")
print(f"✓ State Code: {STATE_CODE}")
if ENTITY_ID:
    print(f"✓ Entity ID:  {ENTITY_ID}")

print("\nFire domains to include — Structure Fire is always included.")
extra = input(
    "Add others? Comma-separated (OUTSIDE, SPECIAL, TRANSPORTATION), "
    "or press Enter for Structure Fire only: "
).strip().upper()

selected_domains = {"FIRE||STRUCTURE_FIRE"}
domain_map = {
    "OUTSIDE": "FIRE||OUTSIDE_FIRE",
    "SPECIAL": "FIRE||SPECIAL_FIRE",
    "TRANSPORTATION": "FIRE||TRANSPORTATION_FIRE",
}
for token in extra.split(","):
    token = token.strip()
    if token in domain_map:
        selected_domains.add(domain_map[token])

print(f"✓ Fire domains: {sorted(selected_domains)}")

start_date_str = input("Start date YYYY-MM-DD (optional, press Enter to skip): ").strip()
end_date_str   = input("End date YYYY-MM-DD (optional, press Enter to skip): ").strip()

api_start_dt = None
api_end_dt = None
if start_date_str:
    api_start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
if end_date_str:
    api_end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").replace(
        hour=23, minute=59, second=59, tzinfo=timezone.utc)

os.environ["NERIS_BASE_URL"]   = "https://api.neris.fsri.org/v1"
os.environ["NERIS_GRANT_TYPE"] = "password"
os.environ["NERIS_USERNAME"]   = NERIS_USERNAME
os.environ["NERIS_PASSWORD"]   = NERIS_PASSWORD

print("\nCreating NERIS API Client...")
client = NerisApiClient()
print("✓ Client created")

print("\n" + "=" * 70)
print("  CHECK YOUR EMAIL FOR THE MFA CODE")
print("=" * 70)

input("Press Enter once you have received the MFA email and are ready to continue...")

client.list_incidents(page_size=1)

print("\n" + "=" * 70)
print("✓✓✓ AUTHENTICATION SUCCESSFUL!")
print("=" * 70)

incident_type_filter = []
for domain in selected_domains:
    incident_type_filter.extend(FIRE_DOMAIN_LEAF_TYPES.get(domain, []))

query_kwargs = {"page_size": 100, "state": STATE_CODE, "incident_types": incident_type_filter}
if ENTITY_ID:
    query_kwargs["neris_id_entity"] = ENTITY_ID
if api_start_dt:
    query_kwargs["call_create_start"] = api_start_dt  # timezone-aware datetime, NOT a string
if api_end_dt:
    query_kwargs["call_create_end"] = api_end_dt      # timezone-aware datetime, NOT a string

print(f"\nQuerying state: {STATE_CODE}" + (f" | entity: {ENTITY_ID}" if ENTITY_ID else " | all departments"))
print(f"Server-side incident_type filter: {len(incident_type_filter)} type(s)")
if api_start_dt:
    print(f"Server-side date filter: call_create >= {api_start_dt.isoformat()}")
if api_end_dt:
    print(f"Server-side date filter: call_create <= {api_end_dt.isoformat()}")

rows = []
cursor = None
page_num = 0

while True:
    page_num += 1
    kwargs = dict(query_kwargs)
    if cursor:
        kwargs["cursor"] = cursor

    print(f"Fetching page {page_num}...")
    try:
        resp = _fetch_page(client, kwargs)
    except Exception as e:
        print(f"✗ Page {page_num} failed after {MAX_RETRIES} attempts: {e}")
        print(f"  Stopping here. Last successful cursor: {cursor!r}")
        break

    incidents = resp.get("incidents", []) if isinstance(resp, dict) else []

    for incident in incidents:
        if not incident_matches(incident, selected_domains):
            continue

        casualties = incident.get("casualty_rescues", []) or []
        if not casualties:
            continue

        base = incident.get("base", {}) or {}
        dispatch = incident.get("dispatch", {}) or {}
        dept_id = base.get("department_neris_id", "")
        type1, type2, type3 = ordered_incident_types(incident)

        for c in casualties:
            casualty_detail = (c.get("casualty") or {}).get("injury_or_noninjury") or {}
            rescue_detail = ((c.get("rescue") or {}).get("ffrescue_or_nonffrescue") or {})

            rows.append({
                "FD NERIS ID":        dept_id,
                "FD Name":            dept_id,  # backfilled after retrieval, below
                "Incident NERIS ID":  incident.get("neris_id", ""),
                "Call Create":        dispatch.get("call_create", ""),
                "Incident Type 1":    type1,
                "Incident Type 2":    type2,
                "Incident Type 3":    type3,
                "Nature of Casualty": casualty_detail.get("type", ""),
                "Cause of Casualty":  casualty_detail.get("cause", ""),
                "Age":                c.get("birth_month_year", ""),
                "Gender":             c.get("gender", ""),
                "Race":               c.get("race", ""),
                "Rescue Type":        rescue_detail.get("type", ""),
            })

    cursor = resp.get("next_cursor") if isinstance(resp, dict) else None

    if page_num % CHECKPOINT_EVERY_N_PAGES == 0 and rows:
        pd.DataFrame(rows).to_csv(CHECKPOINT_PATH, index=False)
        print(f"  💾 Checkpoint saved ({len(rows)} rows so far) → {CHECKPOINT_PATH}")

    if not cursor:
        break

    time.sleep(REQUEST_DELAY_SECONDS)

print(f"\n✓ Extracted {len(rows)} casualty row(s) across matching incidents")

dept_ids_needed = {r["FD NERIS ID"] for r in rows if r["FD NERIS ID"]}
dept_name_lookup = {}
try:
    entities_resp = client.list_entities(page_size=100)
    entities = entities_resp.get("entities", []) if isinstance(entities_resp, dict) else []
    for e in entities:
        nid = e.get("neris_id")
        if nid in dept_ids_needed:
            dept_name_lookup[nid] = e.get("name")
    print(f"✓ Resolved {len(dept_name_lookup)} of {len(dept_ids_needed)} department name(s)")
except Exception as e:
    print(f"⚠ Could not load department name lookup ({e}) — FD Name will fall back to NERIS ID")

for r in rows:
    r["FD Name"] = dept_name_lookup.get(r["FD NERIS ID"], r["FD NERIS ID"])

columns = [
    "FD NERIS ID", "FD Name", "Incident NERIS ID", "Call Create",
    "Incident Type 1", "Incident Type 2", "Incident Type 3",
    "Nature of Casualty", "Cause of Casualty", "Age", "Gender", "Race", "Rescue Type",
]
df = pd.DataFrame(rows, columns=columns)

today_str = datetime.now().strftime("%Y-%m-%d")
output_path = f"Fire Casualties {today_str}.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Casualties")

    if OPENPYXL_AVAILABLE:
        ws = writer.sheets["Casualties"]
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()]
            ) if len(df) else len(str(col_name))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

print(f"\n✓ Saved: {output_path}")
