"""
NERIS Incident Export Tool — States .py
------------------------------------
Fetches incidents for all departments in a state from the NERIS API and
exports them to a formatted Excel file with Overview, Raw Data, and (if
applicable) Casualty & Rescue sheets. 

Note that if a column has no data throughout the entire dataset, it will be dropped from the Excel file output. 

Usage:
    python neris_incident_export_state.py

Requirements (auto-installed if missing):
    pip install pandas openpyxl neris-api-client
"""

import sys
import subprocess
import os
import copy
import re
import traceback
from datetime import datetime, timedelta, timezone


# ─────────────────────────────────────────────
# 1. Install dependencies and NERIS API Client
# ─────────────────────────────────────────────

def ensure_dependencies():
    def pip_install(*packages):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *packages, "--quiet"],
            check=True
        )

    try:
        import pandas   # noqa
        import openpyxl # noqa
    except ImportError:
        print("Installing pandas and openpyxl...")
        pip_install("pandas", "openpyxl")
        print("✓ pandas and openpyxl installed")

    try:
        from neris_api_client import NerisApiClient  # noqa
        print("✓ Dependencies ready")
    except ImportError:
        print("Installing neris-api-client...")
        pip_install("neris-api-client")
        print("✓ neris-api-client installed")


# ─────────────────────────────────────────────
# 2. Define User Inputs
# ─────────────────────────────────────────────

DATE_RANGE_OPTIONS = [
    "All Records",
    "Occurred Today",
    "Occurred Yesterday",
    "Occurred Last 7 Days",
    "Occurred Last 30 Days",
    "Occurred Last 90 Days",
    "Occurred in 2025",
    "Custom Date Range",
]


def prompt_config():
    print("\n" + "=" * 60)
    print("  NERIS Incident Export Tool — State")
    print("=" * 60)

    print("\n── Credentials ──────────────────────────────────")
    username = input("NERIS Email: ").strip()
    print("NERIS Password (note: characters will be visible):")
    password = input("> ").strip()

    print("\n── Query Parameters ─────────────────────────────")
    state_code = input("State Code (e.g. MI, VA, CA): ").strip().upper()

    print("\nDate Range Options:")
    for i, opt in enumerate(DATE_RANGE_OPTIONS, 1):
        print(f"  {i}. {opt}")
    while True:
        choice = input("Select option [1]: ").strip() or "1"
        if choice.isdigit() and 1 <= int(choice) <= len(DATE_RANGE_OPTIONS):
            range_type = DATE_RANGE_OPTIONS[int(choice) - 1]
            break
        print("  Invalid selection, try again.")

    start_date = end_date = None
    if range_type == "Custom Date Range":
        while True:
            s = input("Start date (YYYY-MM-DD): ").strip()
            try:
                start_date = datetime.strptime(s, "%Y-%m-%d").date()
                break
            except ValueError:
                print("  Invalid format, use YYYY-MM-DD.")
        while True:
            e = input("End date (YYYY-MM-DD): ").strip()
            try:
                end_date = datetime.strptime(e, "%Y-%m-%d").date()
                break
            except ValueError:
                print("  Invalid format, use YYYY-MM-DD.")

    return username, password, state_code, range_type, start_date, end_date


# ─────────────────────────────────────────────
# 3. DATE RANGE CALCULATIONS
# ─────────────────────────────────────────────

def calculate_date_range(range_type, start_date=None, end_date=None):
    """Return (call_create_start, call_create_end) as UTC-aware datetimes."""
    now   = datetime.now(tz=timezone.utc)
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)

    if range_type == "All Records":
        return None, None
    elif range_type == "Occurred Today":
        return today, None
    elif range_type == "Occurred Yesterday":
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday.replace(hour=23, minute=59, second=59)
    elif range_type == "Occurred Last 7 Days":
        return today - timedelta(days=7), None
    elif range_type == "Occurred Last 30 Days":
        return today - timedelta(days=30), None
    elif range_type == "Occurred Last 90 Days":
        return today - timedelta(days=90), None
    elif range_type.startswith("Occurred in "):
        year = int(range_type.split()[-1])
        return (
            datetime(year, 1,  1,  0,  0,  0, tzinfo=timezone.utc),
            datetime(year, 12, 31, 23, 59, 59, tzinfo=timezone.utc),
        )
    elif range_type == "Custom Date Range" and start_date and end_date:
        start_dt = datetime.combine(start_date, datetime.min.time(), tzinfo=timezone.utc)
        end_dt   = datetime.combine(
            end_date, datetime.max.time().replace(microsecond=0), tzinfo=timezone.utc
        )
        return start_dt, end_dt

    return None, None


# ─────────────────────────────────────────────
# 4. API AUTHENTICATION
# ─────────────────────────────────────────────

def authenticate(username, password):
    from neris_api_client import NerisApiClient, Config

    print("\nConnecting to NERIS API...")
    client = NerisApiClient(Config(
        base_url="https://api.neris.fsri.org/v1",
        grant_type="password",
        username=username,
        password=password,
    ))

    print("\n✓ Authentication successful!")
    return client


# ─────────────────────────────────────────────
# 5. INCIDENT RETRIEVAL
# ─────────────────────────────────────────────

def get_state_incidents(client, state_code,
                        call_create_start=None, call_create_end=None,
                        page_size=100):
    all_incidents = []
    next_cursor   = None
    page_count    = 0

    print(f"\nFetching incidents for state: {state_code}")
    if call_create_start:
        print(f"  call_create >= {call_create_start:%Y-%m-%d %H:%M:%S}")
    if call_create_end:
        print(f"  call_create <= {call_create_end:%Y-%m-%d %H:%M:%S}")
    if not call_create_start and not call_create_end:
        print("  No date filter — retrieving all records")

    while True:
        page_count += 1
        print(f"  Page {page_count}... ", end="", flush=True)

        kwargs = {"state": state_code, "page_size": page_size}
        if call_create_start: kwargs["call_create_start"] = call_create_start
        if call_create_end:   kwargs["call_create_end"]   = call_create_end
        if next_cursor:       kwargs["cursor"]             = next_cursor

        try:
            res = client.list_incidents(**kwargs)
            if not isinstance(res, dict):
                res = res.json()
        except Exception as e:
            print(f"\n✗ Error on page {page_count}: {e}")
            traceback.print_exc()
            break

        incidents = res.get("incidents", [])
        if not incidents:
            print("empty page — done.")
            break

        all_incidents.extend(incidents)
        print(f"{len(incidents)} fetched  (running total: {len(all_incidents)})")

        next_cursor = res.get("next_cursor")
        if not next_cursor:
            print("  No more pages.")
            break

    print(f"\n{'='*50}")
    print(f"Total incidents retrieved: {len(all_incidents)}")
    print(f"{'='*50}")
    return all_incidents


# ─────────────────────────────────────────────
# 6. EXCEL EXPORT
# ─────────────────────────────────────────────

def export_to_excel(incidents, client, state_code):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not incidents:
        print("⚠ No incidents to export.")
        return None

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"NERIS Incidents {state_code} {date_str}.xlsx"

    # Build department name cache for all unique dept IDs in the results
    dept_ids = {
        (inc.get("base") or {}).get("department_neris_id")
        for inc in incidents
    } - {None, ""}

    print(f"\nLooking up {len(dept_ids)} department names...")
    dept_names = {}
    for i, did in enumerate(dept_ids, 1):
        try:
            entity = client.get_entity(did)
            dept_names[did] = entity.get("name", "")
        except Exception:
            dept_names[did] = ""
        if i % 25 == 0:
            print(f"  {i}/{len(dept_ids)} departments looked up...")
    print("✓ Department lookup complete")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def extract_list_field(lst, key):
        if not lst or not isinstance(lst, list):
            return ""
        return "||".join(
            str(item.get(key, ""))
            for item in lst
            if isinstance(item, dict) and item.get(key)
        )

    def join_list(lst):
        if not lst or not isinstance(lst, list):
            return ""
        return "||".join(str(x) for x in lst if x)

    def get_coordinates_from_point(point_obj):
        if not point_obj or not isinstance(point_obj, dict):
            return None, None
        geometry = point_obj.get("geometry", {})
        if geometry and isinstance(geometry, dict):
            coords = geometry.get("coordinates", [])
            if coords and len(coords) >= 2:
                return coords[1], coords[0]   # GeoJSON is [lon, lat]
        return None, None

    def get_location_data(loc_dict, point_obj=None):
        if not loc_dict or not isinstance(loc_dict, dict):
            return {}
        lat, lon = get_coordinates_from_point(point_obj)
        return {
            "street_number": loc_dict.get("number", ""),
            "street_prefix": loc_dict.get("street_prefix_direction", ""),
            "street_name":   loc_dict.get("street", ""),
            "street_type":   loc_dict.get("street_postfix", ""),
            "street_suffix": loc_dict.get("street_postfix_direction", ""),
            "apartment":     loc_dict.get("unit_value", ""),
            "city":          loc_dict.get("incorporated_municipality", ""),
            "state":         loc_dict.get("state", ""),
            "zip_code":      loc_dict.get("postal_code", ""),
            "county":        loc_dict.get("county", ""),
            "latitude":      lat if lat is not None else "",
            "longitude":     lon if lon is not None else "",
            "census_tract":  "",
            "census_block":  "",
            "parcel_id":     "",
        }

    ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

    def sanitize(val):
        return ILLEGAL_CHARS_RE.sub("", val) if isinstance(val, str) else val

    # ── Overview sheet ────────────────────────────────────────────────────────

    overview_rows = []
    for inc in incidents:
        base       = inc.get("base", {}) or {}
        dispatch   = inc.get("dispatch", {}) or {}
        status_obj = inc.get("incident_status", {}) or {}
        dept_id    = base.get("department_neris_id", "")

        inc_types = inc.get("incident_types", []) or []
        type_list = [
            it.get("type", "")
            for it in inc_types
            if isinstance(it, dict) and it.get("type")
        ]

        overview_rows.append({
            "Incident NERIS ID":    inc.get("neris_id", ""),
            "Department/Entity ID": dept_id,
            "Department Name":      dept_names.get(dept_id, ""),
            "Incident Number":      base.get("incident_number", ""),
            "Call Date":            dispatch.get("call_create", ""),
            "Last Modified":        inc.get("last_modified", ""),
            "Incident Status":      status_obj.get("status", ""),
            "First Incident Type":  type_list[0] if len(type_list) > 0 else "",
            "Second Incident Type": type_list[1] if len(type_list) > 1 else "",
            "Third Incident Type":  type_list[2] if len(type_list) > 2 else "",
            "People Present":       base.get("people_present", ""),
            "Animals Rescued":      base.get("animals_rescued", ""),
            "Outcome Narrative":    base.get("outcome_narrative", ""),
            "Displacement Count":   base.get("displacement_count", ""),
            "Submitter Type":       inc.get("submitter_account_type", ""),
        })

    # ── Raw Data + Casualty Detail ────────────────────────────────────────────

    casualty_detail_rows = []
    processed = []

    for inc in copy.deepcopy(incidents):
        base     = inc.get("base", {}) or {}
        dispatch = inc.get("dispatch", {}) or {}
        dept_id  = base.get("department_neris_id", "")

        inc["department_name"]      = dept_names.get(dept_id, "")
        inc["department_neris_id"]  = dept_id
        inc["incident_number"]      = base.get("incident_number", "")
        inc["department_time_zone"] = (inc.get("department", {}) or {}).get("time_zone", "")

        inc_types = inc.get("incident_types", []) or []
        inc["primary_incident_types"] = "||".join(
            it.get("type", "") for it in inc_types
            if isinstance(it, dict) and it.get("primary")
        )
        inc["other_incident_types"] = "||".join(
            it.get("type", "") for it in inc_types
            if isinstance(it, dict) and not it.get("primary")
        )
        inc.pop("incident_types", None)

        inc["special_modifiers"] = extract_list_field(inc.get("special_modifiers", []), "type")

        aids = inc.get("aids", []) or []
        inc["aid_types"]          = extract_list_field(aids, "aid_type")
        inc["aid_directions"]     = extract_list_field(aids, "aid_direction")
        inc["aid_department_ids"] = extract_list_field(aids, "department_neris_id")
        inc.pop("aids", None)

        inc["nonfd_aid_types"] = extract_list_field(inc.get("nonfd_aids", []), "type")
        inc.pop("nonfd_aids", None)

        units = inc.get("unit_responses", []) or []
        inc["unit_ids"]            = extract_list_field(units, "reported_unit_id")
        inc["unit_neris_ids"]      = extract_list_field(units, "unit_neris_id")
        inc["total_units"]         = len(units)
        inc["total_staffing"]      = sum(
            u.get("staffing", 0) or 0 for u in units if isinstance(u, dict)
        )
        inc["unit_response_modes"] = extract_list_field(units, "response_mode")
        inc.pop("unit_responses", None)

        disp_units = dispatch.get("unit_responses", []) or []
        dispatch["disp_unit_ids"]       = extract_list_field(disp_units, "reported_unit_id")
        dispatch["disp_total_units"]    = len(disp_units)
        dispatch["disp_total_staffing"] = sum(
            u.get("staffing", 0) or 0 for u in disp_units if isinstance(u, dict)
        )
        dispatch.pop("unit_responses", None)

        comments = dispatch.get("comments", []) or []
        dispatch["comments_text"] = extract_list_field(comments, "comment")
        dispatch.pop("comments", None)

        base["displacement_causes"] = join_list(base.get("displacement_causes", []))

        # Location consolidation
        base_loc       = base.get("location", {}) or {}
        base_geo       = base_loc.get("geocoded_location", {}) or {}
        base_geo_point = base_geo.get("point", {}) if base_geo else {}
        disp_loc       = dispatch.get("location", {}) or {}
        disp_geo       = disp_loc.get("geocoded_location", {}) or {}
        disp_geo_point = disp_geo.get("point", {}) if disp_geo else {}

        location_source = ""
        final_location  = {}

        if base_geo and any(base_geo.values()):
            final_location  = get_location_data(base_geo, base_geo_point)
            location_source = "base_location_geocoded"
        elif base_loc and any(base_loc.values()):
            final_location  = get_location_data(base_loc)
            location_source = "base_location"
        elif disp_geo and any(disp_geo.values()):
            final_location  = get_location_data(disp_geo, disp_geo_point)
            location_source = "dispatch_location_geocoded"
        elif disp_loc and any(disp_loc.values()):
            final_location  = get_location_data(disp_loc)
            location_source = "dispatch_location"

        inc["location_source"] = location_source
        for field, value in final_location.items():
            inc[f"location_{field}"] = value
        base.pop("location", None)
        dispatch.pop("location", None)

        exposures = inc.get("exposures", []) or []
        inc["exposure_count"]        = len(exposures)
        inc["exposure_damage_types"] = extract_list_field(exposures, "damage_type")
        inc.pop("exposures", None)

        casualties = inc.get("casualty_rescues", []) or []
        inc["casualty_count"]   = len(casualties)
        inc["casualty_types"]   = extract_list_field(casualties, "type")
        inc["casualty_ranks"]   = extract_list_field(casualties, "rank")
        inc["casualty_genders"] = extract_list_field(casualties, "gender")

        neris_id = inc.get("neris_id", "")
        for cas in casualties:
            if not isinstance(cas, dict):
                continue
            casualty_obj   = cas.get("casualty", {}) or {}
            injury         = casualty_obj.get("injury_or_noninjury", {}) or {}
            ff_details     = injury.get("ff_injury_details", {}) or {}
            rescue_obj     = cas.get("rescue", {}) or {}
            ffrescue       = rescue_obj.get("ffrescue_or_nonffrescue", {}) or {}
            removal        = ffrescue.get("removal_or_nonremoval", {}) or {}
            fire_removal   = removal.get("fire_removal", {}) or {}
            presence_known = rescue_obj.get("presence_known", {}) or {}
            mayday_obj     = rescue_obj.get("mayday", {}) or {}

            casualty_detail_rows.append({
                "Incident NERIS ID":    neris_id,
                "Department/Entity ID": dept_id,
                "Department Name":      dept_names.get(dept_id, ""),
                "Call Date":            dispatch.get("call_create", ""),
                "Type":                 cas.get("type", ""),
                "Rank":                 cas.get("rank", ""),
                "Years of Service":     cas.get("years_of_service", ""),
                "Birth Month Year":     cas.get("birth_month_year", ""),
                "Gender":               cas.get("gender", ""),
                "Race":                 cas.get("race", ""),
                "Injury Type":          injury.get("type", ""),
                "Injury Cause":         injury.get("cause", ""),
                "FF Unit NERIS ID":     ff_details.get("unit_neris_id", ""),
                "FF Reported Unit ID":  ff_details.get("reported_unit_id", ""),
                "FF Unit Continuity":   ff_details.get("unit_continuity", ""),
                "FF Incident Command":  ff_details.get("incident_command", ""),
                "FF Job Classification":ff_details.get("job_classification", ""),
                "FF Duty Type":         ff_details.get("duty_type", ""),
                "FF Action Type":       ff_details.get("action_type", ""),
                "FF Incident Stage":    ff_details.get("incident_stage", ""),
                "FF PPE Items":         join_list(ff_details.get("ppe_items", [])),
                "Rescue Type":          ffrescue.get("type", ""),
                "Rescue Actions":       join_list(ffrescue.get("actions", [])),
                "Rescue Impediments":   join_list(ffrescue.get("impediments", [])),
                "Removal Type":         removal.get("type", ""),
                "Removal Room Type":    removal.get("room_type", ""),
                "Removal Elevation Type": removal.get("elevation_type", ""),
                "Rescue Path Type":     removal.get("rescue_path_type", ""),
                "Gas Isolation":        removal.get("gas_isolation", ""),
                "Fire Removal Relative Suppression Time": fire_removal.get("relative_suppression_time", ""),
                "Presence Known Type":  presence_known.get("presence_known_type", ""),
                "Mayday":               mayday_obj.get("mayday", ""),
                "RIT Activated":        mayday_obj.get("rit_activated", ""),
                "Mayday Relative Suppression Time": mayday_obj.get("relative_suppression_time", ""),
            })

        inc.pop("casualty_rescues", None)

        med = inc.get("medical_details", []) or []
        inc["med_report_ids"]             = extract_list_field(med, "patient_care_report_id")
        inc["med_evaluations"]            = extract_list_field(med, "patient_care_evaluation")
        inc["med_patient_statuses"]       = extract_list_field(med, "patient_status")
        inc["med_transport_dispositions"] = extract_list_field(med, "transport_disposition")
        inc.pop("medical_details", None)

        elec = inc.get("electric_hazards", []) or []
        inc["electric_hazard_types"]         = extract_list_field(elec, "type")
        inc["electric_hazard_source_target"] = extract_list_field(elec, "source_or_target")
        inc.pop("electric_hazards", None)

        inc["powergen_hazard_count"] = len(inc.get("powergen_hazards") or [])
        inc.pop("powergen_hazards", None)

        fire = inc.get("fire_detail", {}) or {}
        fire["investigation_types"]    = join_list(fire.get("investigation_types", []))
        fire["suppression_appliances"] = join_list(fire.get("suppression_appliances", []))

        hazsit = inc.get("hazsit_detail", {}) or {}
        chems  = hazsit.get("chemicals", []) or []
        hazsit["chemical_names"]       = extract_list_field(chems, "name")
        hazsit["chemical_dot_classes"] = extract_list_field(chems, "dot_class")
        hazsit.pop("chemicals", None)

        for alarm_key in ("smoke_alarm", "fire_alarm", "other_alarm"):
            presence = (inc.get(alarm_key, {}) or {}).get("presence", {}) or {}
            presence["alarm_types"] = join_list(presence.get("alarm_types", []))

        fs_presence = (inc.get("fire_suppression", {}) or {}).get("presence", {}) or {}
        supp_types  = fs_presence.get("suppression_types", []) or []
        fs_presence["suppression_types"] = (
            extract_list_field(supp_types, "type")
            if supp_types and isinstance(supp_types[0], dict)
            else join_list(supp_types)
        )

        cfs_presence = (inc.get("cooking_fire_suppression", {}) or {}).get("presence", {}) or {}
        cfs_presence["suppression_types"] = join_list(cfs_presence.get("suppression_types", []))

        at = inc.get("actions_tactics", {}) or {}
        an = at.get("action_noaction", {}) or {}
        an["actions"] = join_list(an.get("actions", []))

        processed.append(inc)

    df_raw = pd.json_normalize(processed, sep="_")

  # If you want to add any of these fields back, you can grab them from this list and add them to the lists above. 
    exclude_cols = {
        "actions_tactics_last_modified", "actions_tactics_neris_uid",
        "base_location_geocoded_location_neris_uid", "base_location_neris_uid",
        "base_location_use_in_use_last_modified", "base_location_use_in_use_neris_uid",
        "base_location_use_neris_uid",
        "census_tract_incident_neris_id", "census_tract_last_modified", "census_tract_neris_uid",
        "cooking_fire_suppression_last_modified", "cooking_fire_suppression_neris_uid",
        "cooking_fire_suppression_presence_last_modified", "cooking_fire_suppression_presence_neris_uid",
        "csst_hazard_last_modified", "csst_hazard_neris_uid",
        "dispatch_last_modified", "dispatch_location_geocoded_location_neris_uid", "dispatch_neris_uid",
        "dispatch_tactic_timestamps_last_modified", "dispatch_tactic_timestamps_neris_uid",
        "fire_alarm_last_modified", "fire_alarm_neris_uid",
        "fire_alarm_presence_last_modified", "fire_alarm_presence_neris_uid",
        "fire_detail_last_modified", "fire_detail_neris_uid",
        "fire_suppression_last_modified", "fire_suppression_neris_uid",
        "fire_suppression_presence_last_modified", "fire_suppression_presence_neris_uid",
        "fire_suppression_presence_operation_type_effectiveness_last_modified",
        "fire_suppression_presence_operation_type_effectiveness_neris_uid",
        "fire_suppression_presence_operation_type_last_modified",
        "fire_suppression_presence_operation_type_neris_uid",
        "hazsit_detail_last_modified", "hazsit_detail_neris_uid",
        "incident_status_last_modified",
        "other_alarm_last_modified", "other_alarm_neris_uid",
        "other_alarm_presence_last_modified", "other_alarm_presence_neris_uid",
        "smoke_alarm_last_modified", "smoke_alarm_neris_uid",
        "smoke_alarm_presence_operation_alerted_failed_other_last_modified",
        "smoke_alarm_presence_operation_alerted_failed_other_neris_uid",
        "smoke_alarm_presence_operation_last_modified", "smoke_alarm_presence_operation_neris_uid",
        "tactic_timestamps_last_modified", "tactic_timestamps_neris_uid",
        "weather_last_modified", "weather_neris_uid",
        "department_time_zone", "base_neris_uid", "base_last_modified",
        "base_department_neris_id", "incident_status_created_by",
        "dispatch_location_neris_uid",
        "tactic_timestamps_command_established", "tactic_timestamps_completed_sizeup",
        "tactic_timestamps_suppression_complete", "tactic_timestamps_primary_search_begin",
        "tactic_timestamps_primary_search_complete", "tactic_timestamps_water_on_fire",
        "tactic_timestamps_fire_under_control", "tactic_timestamps_fire_knocked_down",
        "weather_incident_neris_id",
        "dispatch_tactic_timestamps_command_established",
        "dispatch_tactic_timestamps_completed_sizeup",
        "dispatch_tactic_timestamps_suppression_complete",
        "dispatch_tactic_timestamps_fire_under_control",
    }
    df_raw = df_raw.drop(columns=[c for c in exclude_cols if c in df_raw.columns])

    def reorder_columns(df):
        first_cols = ["department_neris_id", "department_name", "neris_id",
                      "neris_uid", "incident_status_status"]
        location_cols = [
            "location_source", "location_street_number", "location_street_prefix",
            "location_street_name", "location_street_type", "location_street_suffix",
            "location_apartment", "location_city", "location_state",
            "location_zip_code", "location_county", "location_latitude",
            "location_longitude", "location_census_tract", "location_census_block",
            "location_parcel_id",
        ]
        hazard_cols = ["electric_hazard_types", "electric_hazard_source_target",
                       "powergen_hazard_count"]

        existing_first    = [c for c in first_cols    if c in df.columns]
        existing_location = [c for c in location_cols if c in df.columns]
        existing_hazard   = [c for c in hazard_cols   if c in df.columns]
        pinned            = set(existing_first + existing_location + existing_hazard + ["last_modified"])
        middle_cols       = [c for c in df.columns if c not in pinned]

        final_order = existing_first + existing_location + middle_cols + existing_hazard
        if "last_modified" in df.columns:
            final_order.append("last_modified")
        return df[final_order]

    df_raw = reorder_columns(df_raw)

    def clean_df(df):
        drop = [
            col for col in df.columns
            if df[col].dropna().pipe(
                lambda s: s.empty or s.apply(
                    lambda x: x == "" or x == [] or str(x) == "[]"
                ).all()
            )
        ]
        return df.drop(columns=drop)

    df_overview   = clean_df(pd.DataFrame(overview_rows).fillna(""))
    df_raw        = clean_df(df_raw)
    df_casualties = (
        clean_df(pd.DataFrame(casualty_detail_rows).fillna(""))
        if casualty_detail_rows else pd.DataFrame()
    )

    print(f"  Overview:          {len(df_overview.columns)} columns")
    print(f"  Raw Data:          {len(df_raw.columns)} columns")
    print(f"  Casualty & Rescue: {len(df_casualties)} rows")

    def write_sheet(wb, df, name):
        ws      = wb.create_sheet(name)
        headers = list(df.columns)
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append([sanitize(row[h]) for h in headers])

        fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font  = Font(color="FFFFFF", bold=True, size=11)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment = fill, font, align

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len    = len(str(header))
            for row_num in range(2, min(ws.max_row + 1, 52)):
                val = ws.cell(row_num, col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)
        ws.freeze_panes = "A2"

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, df_overview, "Overview")
    write_sheet(wb, df_raw,      "Raw Data")
    if not df_casualties.empty:
        write_sheet(wb, df_casualties, "Casualty & Rescue")

    wb.save(filename)

    sheets = "Overview, Raw Data" + (", Casualty & Rescue" if not df_casualties.empty else "")
    print(f"\n✓ Exported: {filename}")
    print(f"  Incidents : {len(incidents)}")
    print(f"  Sheets    : {sheets}")
    return filename


# ─────────────────────────────────────────────
# 7. MAIN
# ─────────────────────────────────────────────

def main():
    ensure_dependencies()

    username, password, state_code, range_type, start_date, end_date = prompt_config()

    if not username or not password:
        sys.exit("✗ Email and password are required.")
    if not state_code:
        sys.exit("✗ State code is required.")

    client = authenticate(username, password)

    start_filter, end_filter = calculate_date_range(range_type, start_date, end_date)

    incidents = get_state_incidents(
        client,
        state_code,
        call_create_start=start_filter,
        call_create_end=end_filter,
    )

    if not incidents:
        print("⚠ No incidents found for the given parameters.")
        return

    print("\nExporting to Excel...")
    export_to_excel(incidents, client, state_code=state_code)

    print("\n" + "=" * 60)
    print("  PROCESS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
