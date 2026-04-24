#!/usr/bin/env python3
"""
NERIS Grant Report — Incident Type Count by Year
-------------------------------------------------
Fetches all incidents for a single department and produces a formatted
Excel report with four tables:
  - All Incidents (totals, multi-type, aid given/received)
  - Incident Count by Discipline
  - Fire Sub-Category
  - Count of Incidents by Number of Incident Disciplines

Usage:
    python neris_grant_report.py

Requirements (auto-installed if missing):
    pip install python-dateutil openpyxl neris-api-client
"""

import sys
import subprocess
import os
from datetime import datetime


# ─────────────────────────────────────────────
# 1. DEPENDENCY INSTALLATION
# ─────────────────────────────────────────────

def ensure_dependencies():
    def pip_install(*packages):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *packages, "--quiet"],
            check=True
        )

    try:
        import openpyxl    # noqa
        from dateutil import parser  # noqa
    except ImportError:
        print("Installing openpyxl and python-dateutil...")
        pip_install("openpyxl", "python-dateutil")
        print("✓ Installed")

    try:
        from neris_api_client import NerisApiClient  # noqa
        print("✓ Dependencies ready")
    except ImportError:
        print("Installing neris-api-client...")
        pip_install("neris-api-client")
        print("✓ neris-api-client installed")


# ─────────────────────────────────────────────
# 2. INCIDENT TYPE → CATEGORY MAPPING
# ─────────────────────────────────────────────

INCIDENT_TYPE_MAP = {
    # FIRE — Outside Fire
    "FIRE||OUTSIDE_FIRE||CONSTRUCTION_WASTE":               ("FIRE", "TRASH / RUBBISH FIRE"),
    "FIRE||OUTSIDE_FIRE||DUMPSTER_OUTDOOR_CONTAINER_FIRE":  ("FIRE", "TRASH / RUBBISH FIRE"),
    "FIRE||OUTSIDE_FIRE||OTHER_OUTSIDE_FIRE":               ("FIRE", "OUTDOOR FIRE"),
    "FIRE||OUTSIDE_FIRE||OUTSIDE_TANK_FIRE":                ("FIRE", "OUTDOOR FIRE"),
    "FIRE||OUTSIDE_FIRE||TRASH_RUBBISH_FIRE":               ("FIRE", "TRASH / RUBBISH FIRE"),
    "FIRE||OUTSIDE_FIRE||UTILITY_INFRASTRUCTURE_FIRE":      ("FIRE", "OUTDOOR FIRE"),
    "FIRE||OUTSIDE_FIRE||VEGETATION_GRASS_FIRE":            ("FIRE", "VEGETATION FIRE"),
    "FIRE||OUTSIDE_FIRE||WILDFIRE_URBAN_INTERFACE":         ("FIRE", "VEGETATION FIRE"),
    "FIRE||OUTSIDE_FIRE||WILDFIRE_WILDLAND":                ("FIRE", "VEGETATION FIRE"),
    # FIRE — Special
    "FIRE||SPECIAL_FIRE||ESS_FIRE":                         ("FIRE", "SPECIAL FIRE"),
    "FIRE||SPECIAL_FIRE||EXPLOSION":                        ("FIRE", "SPECIAL FIRE"),
    "FIRE||SPECIAL_FIRE||INFRASTRUCTURE_FIRE":              ("FIRE", "SPECIAL FIRE"),
    # FIRE — Structure
    "FIRE||STRUCTURE_FIRE||CHIMNEY_FIRE":                   ("FIRE", "STRUCTURE FIRE"),
    "FIRE||STRUCTURE_FIRE||CONFINED_COOKING_APPLIANCE_FIRE":("FIRE", "STRUCTURE FIRE"),
    "FIRE||STRUCTURE_FIRE||ROOM_AND_CONTENTS_FIRE":         ("FIRE", "STRUCTURE FIRE"),
    "FIRE||STRUCTURE_FIRE||STRUCTURAL_INVOLVEMENT_FIRE":    ("FIRE", "STRUCTURE FIRE"),
    # FIRE — Transportation
    "FIRE||TRANSPORTATION_FIRE||AIRCRAFT_FIRE":                         ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||BOAT_PERSONAL_WATERCRAFT_BARGE_FIRE":   ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||POWERED_MOBILITY_DEVICE_FIRE":          ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||TRAIN_RAIL_FIRE":                       ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_COMMERCIAL":               ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_FOOD_TRUCK":               ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_PASSENGER":                ("FIRE", "TRANSPORTATION FIRE"),
    "FIRE||TRANSPORTATION_FIRE||VEHICLE_FIRE_RV":                       ("FIRE", "TRANSPORTATION FIRE"),
    # HAZSIT
    "HAZSIT||HAZARDOUS_MATERIALS||BIOLOGICAL_RELEASE_INCIDENT":         ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||CARBON_MONOXIDE_RELEASE":             ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||FUEL_SPILL_ODOR":                     ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||GAS_LEAK_ODOR":                       ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||HAZMAT_RELEASE_FACILITY":             ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||HAZMAT_RELEASE_TRANSPORT":            ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARDOUS_MATERIALS||RADIOACTIVE_RELEASE_INCIDENT":        ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARD_NONCHEM||BOMB_THREAT_RESPONSE_SUSPICIOUS_PACKAGE":  ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARD_NONCHEM||ELEC_HAZARD_SHORT_CIRCUIT":                ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARD_NONCHEM||ELEC_POWER_LINE_DOWN_ARCHING_MALFUNC":     ("HAZARDOUS SITUATION", None),
    "HAZSIT||HAZARD_NONCHEM||MOTOR_VEHICLE_COLLISION":                  ("HAZARDOUS SITUATION", None),
    "HAZSIT||INVESTIGATION||ODOR":                                      ("HAZARDOUS SITUATION", None),
    "HAZSIT||INVESTIGATION||SMOKE_INVESTIGATION":                       ("HAZARDOUS SITUATION", None),
    "HAZSIT||OVERPRESSURE||NO_RUPTURE":                                 ("HAZARDOUS SITUATION", None),
    "HAZSIT||OVERPRESSURE||RUPTURE_WITHOUT_FIRE":                       ("HAZARDOUS SITUATION", None),
    # LAW ENFORCE
    "LAWENFORCE": ("PUBLIC SERVICE", None),
    # MEDICAL / EMS
    "MEDICAL||ILLNESS":                                         ("EMS", None),
    "MEDICAL||ILLNESS||ABDOMINAL_PAIN":                         ("EMS", None),
    "MEDICAL||ILLNESS||ALLERGIC_REACTION_STINGS":               ("EMS", None),
    "MEDICAL||ILLNESS||ALTERED_MENTAL_STATUS":                  ("EMS", None),
    "MEDICAL||ILLNESS||BACK_PAIN_NON_TRAUMA":                   ("EMS", None),
    "MEDICAL||ILLNESS||BREATHING_PROBLEMS":                     ("EMS", None),
    "MEDICAL||ILLNESS||CARDIAC_ARREST":                         ("EMS", None),
    "MEDICAL||ILLNESS||CHEST_PAIN_NON_TRAUMA":                  ("EMS", None),
    "MEDICAL||ILLNESS||CONVULSIONS_SEIZURES":                   ("EMS", None),
    "MEDICAL||ILLNESS||DIABETIC_PROBLEMS":                      ("EMS", None),
    "MEDICAL||ILLNESS||HEADACHE":                               ("EMS", None),
    "MEDICAL||ILLNESS||HEART_PROBLEMS":                         ("EMS", None),
    "MEDICAL||ILLNESS||NAUSEA_VOMITING":                        ("EMS", None),
    "MEDICAL||ILLNESS||NO_APPROPRIATE_CHOICE":                  ("EMS", None),
    "MEDICAL||ILLNESS||OVERDOSE":                               ("EMS", None),
    "MEDICAL||ILLNESS||PANDEMIC_EPIDEMIC_OUTBREAK":             ("EMS", None),
    "MEDICAL||ILLNESS||PREGNANCY_CHILDBIRTH":                   ("EMS", None),
    "MEDICAL||ILLNESS||PSYCHOLOGICAL_BEHAVIOR_ISSUES":          ("EMS", None),
    "MEDICAL||ILLNESS||SICK_CASE":                              ("EMS", None),
    "MEDICAL||ILLNESS||STROKE_CVA":                             ("EMS", None),
    "MEDICAL||ILLNESS||UNCONSCIOUS_VICTIM":                     ("EMS", None),
    "MEDICAL||ILLNESS||UNKNOWN_PROBLEM":                        ("EMS", None),
    "MEDICAL||ILLNESS||WELL_PERSON_CHECK":                      ("EMS", None),
    "MEDICAL||INJURY":                                          ("EMS", None),
    "MEDICAL||INJURY||ANIMAL_BITES":                            ("EMS", None),
    "MEDICAL||INJURY||ASSAULT":                                 ("EMS", None),
    "MEDICAL||INJURY||BURNS_EXPLOSION":                         ("EMS", None),
    "MEDICAL||INJURY||CARBON_MONOXIDE_OTHER_INHALATION_INJURY": ("EMS", None),
    "MEDICAL||INJURY||CHOKING":                                 ("EMS", None),
    "MEDICAL||INJURY||DROWNING_DIVING_SCUBA_ACCIDENT":          ("EMS", None),
    "MEDICAL||INJURY||ELECTROCUTION":                           ("EMS", None),
    "MEDICAL||INJURY||EYE_TRAUMA":                              ("EMS", None),
    "MEDICAL||INJURY||FALL":                                    ("EMS", None),
    "MEDICAL||INJURY||GUNSHOT_WOUND":                           ("EMS", None),
    "MEDICAL||INJURY||HEAT_COLD_EXPOSURE":                      ("EMS", None),
    "MEDICAL||INJURY||HEMORRHAGE_LACERATION":                   ("EMS", None),
    "MEDICAL||INJURY||INDUSTRIAL_INACCESSIBLE_ENTRAPMENT":      ("EMS", None),
    "MEDICAL||INJURY||MOTOR_VEHICLE_COLLISION":                 ("EMS", None),
    "MEDICAL||INJURY||OTHER_TRAUMATIC_INJURY":                  ("EMS", None),
    "MEDICAL||INJURY||POISONING":                               ("EMS", None),
    "MEDICAL||INJURY||STAB_PENETRATING_TRAUMA":                 ("EMS", None),
    "MEDICAL||OTHER||AIRMEDICAL_TRANSPORT":                     ("EMS", None),
    "MEDICAL||OTHER||COMMUNITY_PUBLIC_HEALTH":                  ("EMS", None),
    "MEDICAL||OTHER||HEALTHCARE_PROFESSIONAL_ADMISSION":        ("EMS", None),
    "MEDICAL||OTHER||INTERCEPT_OTHER_UNIT":                     ("EMS", None),
    "MEDICAL||OTHER||MEDICAL_ALARM":                            ("EMS", None),
    "MEDICAL||OTHER||STANDBY_REQUEST":                          ("EMS", None),
    "MEDICAL||OTHER||TRANSFER_INTERFACILITY":                   ("EMS", None),
    # FALSE ALARM / NO EMERGENCY
    "NOEMERG||CANCELLED":                                       ("FALSE ALARM", None),
    "NOEMERG||FALSE_ALARM||ACCIDENTAL_ALARM":                   ("FALSE ALARM", None),
    "NOEMERG||FALSE_ALARM||BOMB_SCARE":                         ("FALSE ALARM", None),
    "NOEMERG||FALSE_ALARM||INTENTIONAL_FALSE_ALARM":            ("FALSE ALARM", None),
    "NOEMERG||FALSE_ALARM||MALFUNCTIONING_ALARM":               ("FALSE ALARM", None),
    "NOEMERG||FALSE_ALARM||OTHER_FALSE_CALL":                   ("FALSE ALARM", None),
    # GOOD INTENT
    "NOEMERG||GOOD_INTENT||CONTROLLED_BURNING_AUTHORIZED":      ("GOOD INTENT", None),
    "NOEMERG||GOOD_INTENT||INVESTIGATE_HAZARDOUS_RELEASE":      ("GOOD INTENT", None),
    "NOEMERG||GOOD_INTENT||NO_INCIDENT_FOUND_LOCATION_ERROR":   ("GOOD INTENT", None),
    "NOEMERG||GOOD_INTENT||SMOKE_FROM_NONHOSTILE_SOURCE":       ("GOOD INTENT", None),
    # PUBLIC SERVICE
    "PUBSERV||ALARMS_NONMED||CO_ALARM":                         ("PUBLIC SERVICE", None),
    "PUBSERV||ALARMS_NONMED||FIRE_ALARM":                       ("PUBLIC SERVICE", None),
    "PUBSERV||ALARMS_NONMED||GAS_ALARM":                        ("PUBLIC SERVICE", None),
    "PUBSERV||ALARMS_NONMED||OTHER_ALARM":                      ("PUBLIC SERVICE", None),
    "PUBSERV||CITIZEN_ASSIST||CITIZEN_ASSIST_SERVICE_CALL":     ("PUBLIC SERVICE", None),
    "PUBSERV||CITIZEN_ASSIST||LIFT_ASSIST":                     ("PUBLIC SERVICE", None),
    "PUBSERV||CITIZEN_ASSIST||LOST_PERSON":                     ("PUBLIC SERVICE", None),
    "PUBSERV||CITIZEN_ASSIST||PERSON_IN_DISTRESS":              ("PUBLIC SERVICE", None),
    "PUBSERV||DISASTER_WEATHER||DAMAGE_ASSESSMENT":             ("PUBLIC SERVICE", None),
    "PUBSERV||DISASTER_WEATHER||WEATHER_RESPONSE":              ("PUBLIC SERVICE", None),
    "PUBSERV||OTHER||DAMAGED_HYDRANT":                          ("PUBLIC SERVICE", None),
    "PUBSERV||OTHER||MOVE_UP":                                  ("PUBLIC SERVICE", None),
    "PUBSERV||OTHER||STANDBY":                                  ("PUBLIC SERVICE", None),
    # RESCUE
    "RESCUE||OUTSIDE||BACKCOUNTRY_RESCUE":                      ("RESCUE", None),
    "RESCUE||OUTSIDE||CONFINED_SPACE_RESCUE":                   ("RESCUE", None),
    "RESCUE||OUTSIDE||EXTRICATION_ENTRAPPED":                   ("RESCUE", None),
    "RESCUE||OUTSIDE||HIGH_ANGLE_RESCUE":                       ("RESCUE", None),
    "RESCUE||OUTSIDE||LIMITED_NO_ACCESS":                       ("RESCUE", None),
    "RESCUE||OUTSIDE||LOW_ANGLE_RESCUE":                        ("RESCUE", None),
    "RESCUE||OUTSIDE||STEEP_ANGLE_RESCUE":                      ("RESCUE", None),
    "RESCUE||OUTSIDE||TRENCH":                                  ("RESCUE", None),
    "RESCUE||STRUCTURE||BUILDING_STRUCTURE_COLLAPSE":           ("RESCUE", None),
    "RESCUE||STRUCTURE||CONFINED_SPACE_RESCUE":                 ("RESCUE", None),
    "RESCUE||STRUCTURE||ELEVATOR_ESCALATOR_RESCUE":             ("RESCUE", None),
    "RESCUE||STRUCTURE||EXTRICATION_ENTRAPPED":                 ("RESCUE", None),
    "RESCUE||TRANSPORTATION||AVIATION_COLLISION_CRASH":         ("RESCUE", None),
    "RESCUE||TRANSPORTATION||AVIATION_STANDBY":                 ("RESCUE", None),
    "RESCUE||TRANSPORTATION||MOTOR_VEHICLE_EXTRICATION_ENTRAPPED": ("RESCUE", None),
    "RESCUE||TRANSPORTATION||TRAIN_RAIL_COLLISION_DERAILMENT":  ("RESCUE", None),
    "RESCUE||WATER||PERSON_IN_WATER_STANDING":                  ("RESCUE", None),
    "RESCUE||WATER||PERSON_IN_WATER_SWIFTWATER":                ("RESCUE", None),
}

TABLE0_ROWS = [
    "Total Incidents",
    "Incidents with Multiple Types",
    "Aid Given",
    "Aid Received",
]
TABLE1_ROWS = ["FIRE", "EMS", "RESCUE", "HAZARDOUS SITUATION", "PUBLIC SERVICE", "GOOD INTENT", "FALSE ALARM"]
TABLE2_ROWS = ["STRUCTURE FIRE", "TRANSPORTATION FIRE", "VEGETATION FIRE",
               "TRASH / RUBBISH FIRE", "OUTDOOR FIRE", "SPECIAL FIRE"]


# ─────────────────────────────────────────────
# 3. USER INPUT
# ─────────────────────────────────────────────

def prompt_config():
    print("\n" + "=" * 60)
    print("  NERIS Grant Report — Incident Type Count by Year")
    print("=" * 60)

    print("\n-- Credentials --")
    username  = input("NERIS Email: ").strip()
    print("NERIS Password (note: characters will be visible):")
    password  = input("> ").strip()
    entity_id = input("NERIS Entity ID (e.g. FD26163151): ").strip()

    if not username or not password or not entity_id:
        sys.exit("✗ Email, password, and Entity ID are all required.")

    return username, password, entity_id


# ─────────────────────────────────────────────
# 4. AUTHENTICATION
# ─────────────────────────────────────────────

def authenticate(username, password):
    os.environ["NERIS_BASE_URL"]   = "https://api.neris.fsri.org/v1"
    os.environ["NERIS_GRANT_TYPE"] = "password"
    os.environ["NERIS_USERNAME"]   = username
    os.environ["NERIS_PASSWORD"]   = password

    from neris_api_client import NerisApiClient

    print("\nConnecting to NERIS API...")
    client = NerisApiClient()

    print("\n" + "=" * 60)
    print("  CHECK YOUR EMAIL FOR THE MFA CODE")
    print("=" * 60)

    # First API call triggers the MFA prompt
    client.list_incidents(page_size=1)

    print("\n✓ Authentication successful!")
    return client


# ─────────────────────────────────────────────
# 5. INCIDENT RETRIEVAL
# ─────────────────────────────────────────────

def fetch_all_incidents(client, entity_id, page_size=100):
    all_incidents = []
    next_cursor   = None
    page          = 0

    print(f"\nFetching incidents for: {entity_id}")
    while True:
        page += 1
        print(f"  Page {page}... ", end="", flush=True)
        try:
            kwargs = {"neris_id_entity": entity_id, "page_size": page_size}
            if next_cursor:
                kwargs["cursor"] = next_cursor
            res = client.list_incidents(**kwargs)
            if not isinstance(res, dict):
                res = res.json()
            incidents = res.get("incidents", [])
            if not incidents:
                print("empty — done.")
                break
            all_incidents.extend(incidents)
            print(f"{len(incidents)} retrieved (total: {len(all_incidents)})")
            next_cursor = res.get("next_cursor")
            if not next_cursor:
                print("  No more pages.")
                break
        except Exception as e:
            import traceback
            print(f"\n✗ Error on page {page}: {e}")
            traceback.print_exc()
            break

    print(f"\n✓ Total incidents fetched: {len(all_incidents)}")
    return all_incidents


# ─────────────────────────────────────────────
# 6. DATA PROCESSING
# ─────────────────────────────────────────────

def get_call_year(incident):
    """Extract the year from dispatch.call_create. Returns None if unavailable."""
    from dateutil import parser as dateparser
    try:
        cc = (incident.get("dispatch") or {}).get("call_create")
        if cc:
            return dateparser.parse(cc).year
    except Exception:
        pass
    return None


def get_incident_types(incident):
    """
    Returns a list of dicts: [{type, position}, ...]
    position 1 = primary, 2/3 = additional (by order, capped at 3).
    """
    inc_types = incident.get("incident_types") or []
    result      = []
    primary_seen = False
    non_primary  = []

    for it in inc_types:
        if not isinstance(it, dict):
            continue
        t = (it.get("type") or "").strip()
        if not t:
            continue
        if it.get("primary") and not primary_seen:
            result.append({"type": t, "position": 1})
            primary_seen = True
        else:
            non_primary.append(t)

    for i, t in enumerate(non_primary, start=2):
        result.append({"type": t, "position": min(i, 3)})

    return result


def build_counts(incidents):
    """
    Returns:
        years      : sorted list of years found in the data
        t0_counts  : {metric: {year: count}}
        t1_counts  : {discipline: {year: count}}
        t2_counts  : {fire_subcategory: {year: count}}
        t3_counts  : {discipline: {year: {1: count, 2: count, 3: count}}}
    """
    years    = set()
    t0_counts = {row: {} for row in TABLE0_ROWS}
    t1_counts = {cat: {} for cat in TABLE1_ROWS}
    t2_counts = {sub: {} for sub in TABLE2_ROWS}
    t3_counts = {cat: {} for cat in TABLE1_ROWS}

    for inc in incidents:
        year = get_call_year(inc)
        if year is None:
            continue
        years.add(year)

        # Table 0
        t0_counts["Total Incidents"][year] = t0_counts["Total Incidents"].get(year, 0) + 1

        if len(inc.get("incident_types") or []) > 1:
            t0_counts["Incidents with Multiple Types"][year] = \
                t0_counts["Incidents with Multiple Types"].get(year, 0) + 1

        aids       = inc.get("aids") or []
        directions = {
            (a.get("aid_direction") or "").upper()
            for a in aids if isinstance(a, dict)
        }
        if "GIVEN" in directions:
            t0_counts["Aid Given"][year] = t0_counts["Aid Given"].get(year, 0) + 1
        if "RECEIVED" in directions:
            t0_counts["Aid Received"][year] = t0_counts["Aid Received"].get(year, 0) + 1

        # Tables 1 / 2 / 3
        types            = get_incident_types(inc)
        total_type_count = min(len(types), 3)
        seen_t1          = set()
        seen_t2          = set()

        for t in types:
            mapping = INCIDENT_TYPE_MAP.get(t["type"])
            if not mapping:
                continue
            cat, subcat = mapping

            if cat not in seen_t1:
                seen_t1.add(cat)
                t1_counts[cat][year] = t1_counts[cat].get(year, 0) + 1

            if subcat and subcat not in seen_t2:
                seen_t2.add(subcat)
                t2_counts[subcat][year] = t2_counts[subcat].get(year, 0) + 1

        for cat in seen_t1:
            if year not in t3_counts[cat]:
                t3_counts[cat][year] = {1: 0, 2: 0, 3: 0}
            t3_counts[cat][year][total_type_count] = \
                t3_counts[cat][year].get(total_type_count, 0) + 1

    return sorted(years), t0_counts, t1_counts, t2_counts, t3_counts


# ─────────────────────────────────────────────
# 7. EXCEL REPORT
# ─────────────────────────────────────────────

def write_report(years, t0_counts, t1_counts, t2_counts, t3_counts,
                 entity_id, dept_name, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="262F68", end_color="262F68", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TOTAL_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    DEPT_FILL   = PatternFill(start_color="9A1E22", end_color="9A1E22", fill_type="solid")
    THIN        = Side(style="thin")
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _hcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_BORDER
        return c

    def _dcell(ws, row, col, value, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = THIN_BORDER
        return c

    def _numcell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN_BORDER
        return c

    wb = Workbook()
    ws = wb.active
    ws.title = "Incident Type Report"

    # Table 3 is the widest: 3 sub-columns per year
    total_cols = 1 + len(years) * 3
    row = 1

    # Department banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    banner           = ws.cell(row=row, column=1, value=f"{dept_name}  |  {entity_id}")
    banner.fill      = DEPT_FILL
    banner.font      = Font(color="FFFFFF", bold=True, size=14)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    row += 1  # blank spacer

    # ── Table 0: All Incidents ────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="All Incidents").font = Font(bold=True, size=13)
    row += 1

    _hcell(ws, row, 1, "Metric")
    for i, yr in enumerate(years, start=2):
        _hcell(ws, row, i, yr)
    row += 1

    for metric in TABLE0_ROWS:
        _dcell(ws, row, 1, metric)
        for i, yr in enumerate(years, start=2):
            _numcell(ws, row, i, t0_counts[metric].get(yr, 0))
        row += 1

    row += 2  # spacer

    # ── Table 1: Incident Count by Discipline ─────────────────────────────────
    ws.cell(row=row, column=1, value="Incident Count by Discipline").font = Font(bold=True, size=13)
    row += 1

    _hcell(ws, row, 1, "Discipline")
    for i, yr in enumerate(years, start=2):
        _hcell(ws, row, i, yr)
    row += 1

    for cat in TABLE1_ROWS:
        _dcell(ws, row, 1, cat)
        for i, yr in enumerate(years, start=2):
            _numcell(ws, row, i, t1_counts[cat].get(yr, 0))
        row += 1

    row += 2  # spacer

    # ── Table 2: Fire Sub-Category ────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Fire Sub-Category").font = Font(bold=True, size=13)
    row += 1

    _hcell(ws, row, 1, "Fire Subcategory")
    for i, yr in enumerate(years, start=2):
        _hcell(ws, row, i, yr)
    row += 1

    for sub in TABLE2_ROWS:
        _dcell(ws, row, 1, sub)
        for i, yr in enumerate(years, start=2):
            _numcell(ws, row, i, t2_counts[sub].get(yr, 0))
        row += 1

    row += 2  # spacer

    # ── Table 3: Count of Incidents by Number of Incident Disciplines ──────────
    ws.cell(row=row, column=1,
            value="Count of Incidents by Number of Incident Disciplines"
            ).font = Font(bold=True, size=13)
    row += 1

    # Year super-header, merged across 3 sub-columns each
    _hcell(ws, row, 1, "")
    col_start = 2
    for yr in years:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_start + 2
        )
        c           = ws.cell(row=row, column=col_start, value=yr)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        for mc in range(col_start, col_start + 3):
            ws.cell(row=row, column=mc).border = THIN_BORDER
        col_start += 3
    row += 1

    # Sub-header: One / Two / Three
    _hcell(ws, row, 1, "")
    col = 2
    for _ in years:
        _hcell(ws, row, col,     "One")
        _hcell(ws, row, col + 1, "Two")
        _hcell(ws, row, col + 2, "Three")
        col += 3
    row += 1

    for cat in TABLE1_ROWS:
        _dcell(ws, row, 1, cat)
        col = 2
        for yr in years:
            yr_data = (t3_counts.get(cat) or {}).get(yr, {})
            _numcell(ws, row, col,     yr_data.get(1, 0))
            _numcell(ws, row, col + 1, yr_data.get(2, 0))
            _numcell(ws, row, col + 2, yr_data.get(3, 0))
            col += 3
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    wb.save(filename)
    print(f"\n✓ Report saved: {filename}")


# ─────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────

def main():
    ensure_dependencies()

    username, password, entity_id = prompt_config()

    client = authenticate(username, password)

    # Department name lookup
    try:
        entity    = client.get_entity(entity_id)
        dept_name = entity.get("name", entity_id) if isinstance(entity, dict) else entity_id
    except Exception:
        dept_name = entity_id
    print(f"\nDepartment: {dept_name} ({entity_id})")

    incidents = fetch_all_incidents(client, entity_id)

    if not incidents:
        print("\n⚠ No incidents found for this entity.")
        return

    print("\nBuilding counts...")
    years, t0_counts, t1_counts, t2_counts, t3_counts = build_counts(incidents)

    if not years:
        print("⚠ No incidents with a valid call_create date found.")
        return

    print(f"  Years found: {', '.join(str(y) for y in years)}")

    date_str      = datetime.now().strftime("%Y-%m-%d")
    safe_name     = "".join(c if c.isalnum() or c in " -_" else "" for c in dept_name).strip()
    desktop       = os.path.join(os.path.expanduser("~"), "Desktop")
    filename      = os.path.join(desktop, f"{safe_name} {date_str}.xlsx")

    write_report(years, t0_counts, t1_counts, t2_counts, t3_counts,
                 entity_id, dept_name, filename)

    print("\n" + "=" * 60)
    print("  PROCESS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
