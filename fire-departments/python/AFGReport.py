#!/usr/bin/env python3
"""
Combined AFG Grant Report — NFIRS + NERIS
This produces a single Excel workbook with two sheets intended for AFG incident summary data:
  Sheet 1: NFIRS  (source: ArcGIS Public Data Release Lite, 2020-2024)
  Sheet 2: NERIS  (source: NERIS API)
"""

import sys
import subprocess
import os
from datetime import datetime

# Dependencies - This pip installs dependent pacakges and the NERIS API client
def ensure_dependencies():
    def pip_install(*pkgs):
        subprocess.run(
            [sys.executable, "-m", "pip", "install", *pkgs, "--quiet"],
            check=True,
        )
    try:
        import openpyxl, requests  # noqa
        from dateutil import parser  # noqa
    except ImportError:
        pip_install("openpyxl", "requests", "python-dateutil")
    try:
        from neris_api_client import NerisApiClient  # noqa
    except ImportError:
        pip_install("neris-api-client")
    print("✓ Dependencies ready")


# Report structure
# This gives both sheets their platform-appropriate names (NFIRS are series, NERIS uses the discipline name)

NFIRS_REPORT_GROUPS = [
    {"section": "Casualties"},
    {"label": "Fire-related civilian fatalities",  "key": "civ_fatal"},
    {"label": "Fire-related civilian injuries",    "key": "civ_inj"},
    {"label": "LOD injuries and deaths",           "key": "lod_total"},

    {"section": "NFIRS Incidents"},
    {"label": "100 series",  "key": "s_fire"},
    {"label": "200 series",  "key": "s_overpressure"},
    {"label": "300 series",  "key": "s_ems_rescue"},
    {"label": "400 series",  "key": "s_hazsit"},
    {"label": "500 series",  "key": "s_pubserv"},
    {"label": "600 series",  "key": "s_good_intent"},
    {"label": "700 series",  "key": "s_false_alarm"},
    {"label": "800 series",  "key": "s_weather"},
    {"label": "900 series",  "key": "s_special"},
    {"label": "Total",       "key": "total"},

    {"section": "Fire Detail"},
    {"label": "Structure fires (111-123)",  "key": "fire_struct"},
    {"label": "Vehicle fires (130-138)",    "key": "fire_vehicle"},
    {"label": "Vegetation fires (140-143)", "key": "fire_veg"},
    {"label": "Acreage for veg fires",      "key": "fire_acres"},

    {"section": "Aid"},
    {"label": "Receive mutual aid",            "key": "aid_recv_mutual"},
    {"label": "Receive auto aid",              "key": "aid_recv_auto"},
    {"label": "Provide mutual aid",            "key": "aid_given_mutual"},
    {"label": "Provide auto aid",              "key": "aid_given_auto"},
    {"label": "Aid incidents that were fires", "key": "aid_fire"},

    {"section": "EMS"},
    {"label": "Motor Vehicle Accident Calls (322-324)",   "key": "ems_mva"},
    {"label": "Extrication from Vehicle Calls (352)",     "key": "ems_extric"},
    {"label": "Rescue Calls (300, 350-381)",              "key": "ems_rescue"},
    {"label": "EMS-BLS Calls",                            "key": "ems_bls"},
    {"label": "EMS-ALS Calls",                            "key": "ems_als"},
    {"label": "EMS-BLS Transport",                        "key": "ems_bls_transport"},
    {"label": "EMS-ALS Transport",                        "key": "ems_als_transport"},
]

NERIS_REPORT_GROUPS = [
    {"section": "Casualties"},
    {"label": "Fire-related civilian fatalities",  "key": "civ_fatal"},
    {"label": "Fire-related civilian injuries",    "key": "civ_inj"},
    {"label": "LOD injuries and deaths",           "key": "lod_total"},

    {"section": "Incidents by Discipline"},
    {"label": "Fire",                       "key": "s_fire"},
    {"label": "Overpressures & Explosions", "key": "s_overpressure"},
    {"label": "EMS & Rescue",               "key": "s_ems_rescue"},
    {"label": "HAZSIT",                     "key": "s_hazsit"},
    {"label": "Public Service",             "key": "s_pubserv"},
    {"label": "Good Intent",               "key": "s_good_intent"},
    {"label": "False Alarm",               "key": "s_false_alarm"},
    {"label": "Severe Weather",            "key": "s_weather"},
    {"label": "Total",                     "key": "total"},

    {"section": "Fire Detail"},
    {"label": "Structure fires",       "key": "fire_struct"},
    {"label": "Vehicle fires",         "key": "fire_vehicle"},
    {"label": "Vegetation fires",      "key": "fire_veg"},
    {"label": "Acreage for veg fires", "key": "fire_acres"},

    {"section": "Aid"},
    {"label": "Received aid",                  "key": "aid_recv"},
    {"label": "Provided aid",                  "key": "aid_given"},
    {"label": "Aid incidents that were fires", "key": "aid_fire"},

    {"section": "EMS"},
    {"label": "Motor Vehicle Accident Calls",   "key": "ems_mva"},
    {"label": "Extrication from Vehicle Calls", "key": "ems_extric"},
    {"label": "Rescue Calls",                   "key": "ems_rescue"},
    {"label": "EMS-BLS Calls",                  "key": "ems_bls"},
    {"label": "EMS-ALS Calls",                  "key": "ems_als"},
    {"label": "EMS-BLS Transport",              "key": "ems_bls_transport"},
    {"label": "EMS-ALS Transport",              "key": "ems_als_transport"},
]

# Union of all keys across both group definitions
ALL_KEYS = list(dict.fromkeys(
    [g["key"] for g in NFIRS_REPORT_GROUPS if "key" in g]
    + [g["key"] for g in NERIS_REPORT_GROUPS if "key" in g]
))


def empty_year_dict():
    return {k: {} for k in ALL_KEYS}


#  NFIRS — ArcGIS NFIRS Public Data Release Lite REST API services
# Will add 2025 when available
NFIRS_FEATURE_SERVICES = [
    {"year": 2020, "url": "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/NFIRS_PDR_Light_Service_2020/FeatureServer/0"},
    {"year": 2021, "url": "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/NFIRS_PDR_Light_2021/FeatureServer/0"},
    {"year": 2022, "url": "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/NFIRS_PDR_Light_Service_2022/FeatureServer/0"},
    {"year": 2023, "url": "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/NFIRS_PDR_Light_Service_2023/FeatureServer/0"},
    {"year": 2024, "url": "https://services.arcgis.com/XG15cJAlne2vxtgt/arcgis/rest/services/NFIRS_PDR_Light_Service_2024/FeatureServer/0"},
]

NFIRS_FIELDS = [
    "INC_TYPE", "AID", "INC_DATE", "FD_NAME",
    "FF_DEATH", "OTH_DEATH", "FF_INJ", "OTH_INJ",
    "ACT_TAK1", "ACT_TAK2", "ACT_TAK3",
    "ACRES_BURN",
]


def fetch_nfirs_records(state, fdid, years=None):
    import requests as req

    if years is None:
        years = [s["year"] for s in NFIRS_FEATURE_SERVICES]

    all_records = []
    for svc in NFIRS_FEATURE_SERVICES:
        if svc["year"] not in years:
            continue
        base_url = svc["url"] + "/query"
        where = f"STATE = '{state}' AND FDID = '{fdid}'"
        offset, page_size, layer_n = 0, 2000, 0
        print(f"  NFIRS {svc['year']}… ", end="", flush=True)

        while True:
            params = {
                "where": where,
                "outFields": ",".join(NFIRS_FIELDS),
                "returnGeometry": "false",
                "resultOffset": offset,
                "resultRecordCount": page_size,
                "orderByFields": "OBJECTID ASC",
                "f": "json",
            }
            resp = req.get(base_url, params=params, timeout=120)
            data = resp.json()
            if "error" in data:
                print(f"error – {data['error'].get('message','?')}")
                break
            feats = data.get("features", [])
            if not feats:
                break
            for f in feats:
                attrs = f.get("attributes", {})
                attrs["_year"] = svc["year"]
                all_records.append(attrs)
            layer_n += len(feats)
            offset += page_size
            if len(feats) < page_size:
                break

        print(f"{layer_n:,}")

    print(f"  ✓ Total NFIRS records: {len(all_records):,}")
    return all_records


def _safe_int(v):
    try:
        return int(v or 0)
    except (ValueError, TypeError):
        return 0


def _safe_float(v):
    try:
        return float(v or 0)
    except (ValueError, TypeError):
        return 0.0


def _inc(d, year, val=1):
    d[year] = d.get(year, 0) + val


def _nfirs_is_fire(inc_type):
    return inc_type.startswith("1")


def _nfirs_act_has(rec, code):
    """Check if any of ACT_TAK1/2/3 equals the given code string."""
    for fld in ("ACT_TAK1", "ACT_TAK2", "ACT_TAK3"):
        v = str(rec.get(fld) or "").strip()
        if v == code:
            return True
    return False


# NFIRS structure-fire codes (NFIRS 111-123)
STRUCT_CODES = {str(c) for c in range(111, 124)} | {"11", "110"}
# Vehicle-fire codes (130-138)
VEH_CODES = {str(c) for c in range(130, 139)} | {"13"}
# Vegetation-fire codes (140-143)
VEG_CODES = {str(c) for c in range(140, 144)} | {"14"}
# MVA codes (322-324)
MVA_CODES = {"322", "323", "324"}
# Extrication code
EXTRIC_CODES = {"352"}
# Rescue codes (300, 350-381) excluding extrication
RESCUE_CODES = (
    {"300"}
    | {str(c) for c in range(350, 382)}
) - EXTRIC_CODES

# Begin building the NFIRS counts/Sums
def build_nfirs_counts(records):
    years_set = set()
    C = empty_year_dict()

    for rec in records:
        year = rec.get("_year")
        if year is None:
            continue
        years_set.add(year)

        it = str(rec.get("INC_TYPE") or "").strip()
        series = it[0] if it else ""

        # Total
        _inc(C["total"], year)

        # Casualties (fire-related only)
        if _nfirs_is_fire(it):
            _inc(C["civ_fatal"], year, _safe_int(rec.get("OTH_DEATH")))
            _inc(C["civ_inj"],   year, _safe_int(rec.get("OTH_INJ")))
            _inc(C["lod_total"], year,
                 _safe_int(rec.get("FF_DEATH")) + _safe_int(rec.get("FF_INJ")))

        # Discipline series
        if series == "1":
            _inc(C["s_fire"], year)
        elif series == "2":
            _inc(C["s_overpressure"], year)
        elif series == "3":
            _inc(C["s_ems_rescue"], year)
        elif series == "4":
            _inc(C["s_hazsit"], year)
        elif series == "5":
            _inc(C["s_pubserv"], year)
        elif series == "6":
            _inc(C["s_good_intent"], year)
        elif series == "7":
            _inc(C["s_false_alarm"], year)
        elif series == "8":
            _inc(C["s_weather"], year)
        elif series == "9":
            _inc(C["s_special"], year)

        # Fire detail
        if it in STRUCT_CODES:
            _inc(C["fire_struct"], year)
        if it in VEH_CODES:
            _inc(C["fire_vehicle"], year)
        if it in VEG_CODES:
            _inc(C["fire_veg"], year)
            _inc(C["fire_acres"], year, _safe_float(rec.get("ACRES_BURN")))

        # Aid
        aid = str(rec.get("AID") or "").strip()
        if aid == "1":
            _inc(C["aid_recv_mutual"], year)
        if aid == "2":
            _inc(C["aid_recv_auto"], year)
        if aid in ("1", "2"):
            _inc(C["aid_recv"], year)
        if aid == "3":
            _inc(C["aid_given_mutual"], year)
        if aid == "4":
            _inc(C["aid_given_auto"], year)
        if aid in ("3", "4"):
            _inc(C["aid_given"], year)
        if aid in ("1", "2", "3", "4") and _nfirs_is_fire(it):
            _inc(C["aid_fire"], year)

        # EMS
        if it in MVA_CODES:
            _inc(C["ems_mva"], year)
        if it in EXTRIC_CODES:
            _inc(C["ems_extric"], year)
        if it in RESCUE_CODES:
            _inc(C["ems_rescue"], year)
# Pulls EMS ALS and BLS codes from actions taken

        is_bls = _nfirs_act_has(rec, "32")
        is_als = _nfirs_act_has(rec, "33")
        is_transport = _nfirs_act_has(rec, "34")

        if is_bls:
            _inc(C["ems_bls"], year)
        if is_als:
            _inc(C["ems_als"], year)
        if is_bls and is_transport:
            _inc(C["ems_bls_transport"], year)
        if is_als and is_transport:
            _inc(C["ems_als_transport"], year)

    return sorted(years_set), C

# NERIS — API Fetch
# Notes: NERIS incidents can have up to 3 types, so we check all types for each incident and count it in a bucket if any type matches the criteria. 
# There will be two total fields: One that sums all incidents, and one that sums only incidents that match a discipline (e.g. FIRE). Your discipline sum will likely exceed your total incident count because some incidents will have multiple disciplines. If an incident has multiple types of the same discipline, they are only counted once.  
# This is not an official NFIRS:NERIS incident type mapping. Rather, these are NERIS types that seem to best match the intent of each NFIRS discipline category for the purpose of AFG grant reporting.

# Types that count for each "series-equivalent" bucket
NERIS_FIRE_PREFIX        = "FIRE||"
NERIS_OVERPRESSURE_TYPES = {
    "HAZSIT||OVERPRESSURE||NO_RUPTURE",
    "HAZSIT||OVERPRESSURE||RUPTURE_WITHOUT_FIRE",
}
NERIS_EMS_PREFIX         = "MEDICAL||"
NERIS_RESCUE_PREFIX      = "RESCUE||"
NERIS_HAZSIT_TYPES = {
    "HAZSIT||HAZARDOUS_MATERIALS||BIOLOGICAL_RELEASE_INCIDENT",
    "HAZSIT||HAZARDOUS_MATERIALS||CARBON_MONOXIDE_RELEASE",
    "HAZSIT||HAZARDOUS_MATERIALS||FUEL_SPILL_ODOR",
    "HAZSIT||HAZARDOUS_MATERIALS||GAS_LEAK_ODOR",
    "HAZSIT||HAZARDOUS_MATERIALS||HAZMAT_RELEASE_FACILITY",
    "HAZSIT||HAZARDOUS_MATERIALS||HAZMAT_RELEASE_TRANSPORT",
    "HAZSIT||HAZARDOUS_MATERIALS||RADIOACTIVE_RELEASE_INCIDENT",
    "HAZSIT||HAZARD_NONCHEM||BOMB_THREAT_RESPONSE_SUSPICIOUS_PACKAGE",
    "HAZSIT||HAZARD_NONCHEM||ELEC_HAZARD_SHORT_CIRCUIT",
    "HAZSIT||HAZARD_NONCHEM||ELEC_POWER_LINE_DOWN_ARCHING_MALFUNC",
    "HAZSIT||HAZARD_NONCHEM||MOTOR_VEHICLE_COLLISION",
    "HAZSIT||INVESTIGATION||ODOR",
    "HAZSIT||INVESTIGATION||SMOKE_INVESTIGATION",
}
NERIS_PUBSERV_PREFIXES = ("PUBSERV||", "LAWENFORCE")
# Excluduing any severe weather from pubserv to be included in its own bin (800 series equivalent)
NERIS_PUBSERV_EXCLUDE  = "PUBSERV||DISASTER_WEATHER"
NERIS_GOOD_INTENT_PREFIX = "NOEMERG||GOOD_INTENT||"
NERIS_FALSE_ALARM_TYPES = {
    "NOEMERG||CANCELLED",
    "NOEMERG||FALSE_ALARM||ACCIDENTAL_ALARM",
    "NOEMERG||FALSE_ALARM||BOMB_SCARE",
    "NOEMERG||FALSE_ALARM||INTENTIONAL_FALSE_ALARM",
    "NOEMERG||FALSE_ALARM||MALFUNCTIONING_ALARM",
    "NOEMERG||FALSE_ALARM||OTHER_FALSE_CALL",
}
NERIS_WEATHER_PREFIX = "PUBSERV||DISASTER_WEATHER"

# Fire detail prefixes
NERIS_STRUCT_PREFIX  = "FIRE||STRUCTURE_FIRE||"
NERIS_TRANSP_PREFIX  = "FIRE||TRANSPORTATION_FIRE||"
# All other outside_fires are excluded since it's veg specific
NERIS_VEG_TYPES = {
    "FIRE||OUTSIDE_FIRE||VEGETATION_GRASS_FIRE",
    "FIRE||OUTSIDE_FIRE||WILDFIRE_URBAN_INTERFACE",
    "FIRE||OUTSIDE_FIRE||WILDFIRE_WILDLAND",
}

# MVA types
NERIS_MVA_TYPES = {
    "HAZSIT||HAZARD_NONCHEM||MOTOR_VEHICLE_COLLISION",
    "MEDICAL||INJURY||MOTOR_VEHICLE_COLLISION",
}
NERIS_EXTRIC_TYPE = "RESCUE||TRANSPORTATION||MOTOR_VEHICLE_EXTRICATION_ENTRAPPED"

# Pulls the ALS/BLS/Transport actions taken codes
NERIS_BLS_ACTION       = "EMERGENCY_MEDICAL_CARE||PROVIDE_BASIC_LIFE_SUPPORT"
NERIS_ALS_ACTION       = "EMERGENCY_MEDICAL_CARE||PROVIDE_ADVANCED_LIFE_SUPPORT"
NERIS_TRANSPORT_ACTION = "EMERGENCY_MEDICAL_CARE||PROVIDE_TRANSPORT"

# Define incident type fields since there are three
def _neris_types(inc):
    """Return list of type strings from incident_types (up to 3)."""
    return [
        (it.get("type") or "").strip()
        for it in (inc.get("incident_types") or [])[:3]
        if isinstance(it, dict) and (it.get("type") or "").strip()
    ]


def _neris_any_type_match(types, predicate):
    """True if any type string satisfies predicate."""
    return any(predicate(t) for t in types)


def _neris_actions(inc):
    """Return the set of action strings."""
    at = inc.get("actions_tactics") or {}
    an = at.get("action_noaction") or {}
    return set(an.get("actions") or [])

# The year of the report is pulled from call_create
def neris_get_year(inc):
    from dateutil import parser as dp
    try:
        cc = (inc.get("dispatch") or {}).get("call_create")
        if cc:
            return dp.parse(cc).year
    except Exception:
        pass
    return None

# Defining the auth for the NERIS API
def neris_authenticate(username, password):
    os.environ["NERIS_BASE_URL"]   = "https://api.neris.fsri.org/v1"
    os.environ["NERIS_GRANT_TYPE"] = "password"
    os.environ["NERIS_USERNAME"]   = username
    os.environ["NERIS_PASSWORD"]   = password

    from neris_api_client import NerisApiClient
    print("\nConnecting to NERIS API …")
    client = NerisApiClient()
    print("\n" + "=" * 60)
    print("  CHECK YOUR EMAIL FOR THE MFA CODE")
    print("=" * 60)
    client.list_incidents(page_size=1)
    print("✓ NERIS auth successful")
    return client

# Pull NERIS data
def fetch_neris_incidents(client, entity_id, page_size=100):
    all_inc = []
    cursor = None
    page = 0
    print(f"\n  Fetching NERIS incidents for {entity_id} …")
    while True:
        page += 1
        print(f"    Page {page} … ", end="", flush=True)
        try:
            kw = {"neris_id_entity": entity_id, "page_size": page_size}
            if cursor:
                kw["cursor"] = cursor
            res = client.list_incidents(**kw)
            if not isinstance(res, dict):
                res = res.json()
            incs = res.get("incidents", [])
            if not incs:
                print("empty — done.")
                break
            all_inc.extend(incs)
            print(f"{len(incs)} (total: {len(all_inc)})")
            cursor = res.get("next_cursor")
            if not cursor:
                break
        except Exception as e:
            print(f"\n  ✗ Error page {page}: {e}")
            break
    print(f"  ✓ Total NERIS incidents: {len(all_inc):,}")
    return all_inc

# Begin compiling the NERIS data
def build_neris_counts(incidents):
    years_set = set()
    C = empty_year_dict()

    for inc in incidents:
        year = neris_get_year(inc)
        if year is None:
            continue
        years_set.add(year)

        types   = _neris_types(inc)
        actions = _neris_actions(inc)

        is_fire = _neris_any_type_match(types, lambda t: t.startswith(NERIS_FIRE_PREFIX))

        # Total
        _inc(C["total"], year)

        # Casualties
        for cr in (inc.get("casualty_rescues") or []):
            if not isinstance(cr, dict):
                continue
            cr_type = (cr.get("type") or "").upper()
            cas = cr.get("casualty") or {}
            inj = cas.get("injury_or_noninjury") or {}
            cas_type = (inj.get("type") or "").upper()

            if is_fire:
                if cr_type != "FF":
                    if "FATAL" in cas_type and "NON" not in cas_type:
                        _inc(C["civ_fatal"], year)
                    elif "NONFATAL" in cas_type:
                        _inc(C["civ_inj"], year)
                else:
                    if "FATAL" in cas_type or "NONFATAL" in cas_type:
                        _inc(C["lod_total"], year)

        # Discipline series
        if is_fire:
            _inc(C["s_fire"], year)

        if _neris_any_type_match(types, lambda t: t in NERIS_OVERPRESSURE_TYPES):
            _inc(C["s_overpressure"], year)

        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_EMS_PREFIX) or t.startswith(NERIS_RESCUE_PREFIX)):
            _inc(C["s_ems_rescue"], year)

        if _neris_any_type_match(types, lambda t: t in NERIS_HAZSIT_TYPES):
            _inc(C["s_hazsit"], year)

        # Public Service (PUBSERV + LAWENFORCE, excluding DISASTER_WEATHER)
        if _neris_any_type_match(types, lambda t: (
            any(t.startswith(p) for p in NERIS_PUBSERV_PREFIXES)
            and not t.startswith(NERIS_PUBSERV_EXCLUDE)
        )):
            _inc(C["s_pubserv"], year)

        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_GOOD_INTENT_PREFIX)):
            _inc(C["s_good_intent"], year)

        if _neris_any_type_match(types, lambda t: t in NERIS_FALSE_ALARM_TYPES):
            _inc(C["s_false_alarm"], year)

        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_WEATHER_PREFIX)):
            _inc(C["s_weather"], year)

        # Fire detail (structure, vehicle, veg)
        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_STRUCT_PREFIX)):
            _inc(C["fire_struct"], year)
        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_TRANSP_PREFIX)):
            _inc(C["fire_vehicle"], year)
        if _neris_any_type_match(types, lambda t: t in NERIS_VEG_TYPES):
            _inc(C["fire_veg"], year)
        # Acreage from fire_detail.location_detail.acres_burned - this is not from Incident Analysis
        fd = inc.get("fire_detail") or {}
        ld = fd.get("location_detail") or {}
        acres = _safe_float(ld.get("acres_burned"))
        if acres:
            _inc(C["fire_acres"], year, acres)

        # Aid
        aids = inc.get("aids") or []
        directions = {
            (a.get("aid_direction") or "").upper()
            for a in aids if isinstance(a, dict)
        }
        if "RECEIVED" in directions:
            _inc(C["aid_recv"], year)
        if "GIVEN" in directions:
            _inc(C["aid_given"], year)
        if ("GIVEN" in directions or "RECEIVED" in directions) and is_fire:
            _inc(C["aid_fire"], year)

        # EMS
        if _neris_any_type_match(types, lambda t: t in NERIS_MVA_TYPES):
            _inc(C["ems_mva"], year)
        if _neris_any_type_match(types, lambda t: t == NERIS_EXTRIC_TYPE):
            _inc(C["ems_extric"], year)
        # Rescue calls: RESCUE discipline EXCLUDING motor-vehicle extrication since they are their own row
        if _neris_any_type_match(types, lambda t: t.startswith(NERIS_RESCUE_PREFIX) and t != NERIS_EXTRIC_TYPE):
            _inc(C["ems_rescue"], year)

        has_bls       = NERIS_BLS_ACTION in actions
        has_als       = NERIS_ALS_ACTION in actions
        has_transport = NERIS_TRANSPORT_ACTION in actions

        if has_bls:
            _inc(C["ems_bls"], year)
        if has_als:
            _inc(C["ems_als"], year)
        if has_bls and has_transport:
            _inc(C["ems_bls_transport"], year)
        if has_als and has_transport:
            _inc(C["ems_als_transport"], year)

    return sorted(years_set), C

# Excel report writer

def write_sheet(ws, years, counts, sheet_label, entity_label, dept_name, report_groups):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="262F68", end_color="262F68", fill_type="solid")
    HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    SECT_FILL   = PatternFill(start_color="3F6F8F", end_color="3F6F8F", fill_type="solid")
    SECT_FONT   = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    DEPT_FILL   = PatternFill(start_color="9A1E22", end_color="9A1E22", fill_type="solid")
    THIN        = Side(style="thin")
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HAIR_BOTTOM = Border(bottom=Side(style="hair", color="E0E0E0"))
    DATA_FONT   = Font(name="Arial", size=11)
    BOLD_FONT   = Font(name="Arial", bold=True, size=11)
    TOTAL_FONT  = Font(name="Arial", bold=True, size=11, color="17324D")

    num_cols = 2 + len(years)  # label + years + total
    row = 1

    # Banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    b = ws.cell(row=row, column=1,
                value=f"{sheet_label}  —  {dept_name}  |  {entity_label}")
    b.fill = DEPT_FILL
    b.font = Font(name="Arial", color="FFFFFF", bold=True, size=14)
    b.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # Sub-line
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    s = ws.cell(row=row, column=1,
                value=f"Generated {datetime.now().strftime('%m/%d/%Y')}")
    s.font = Font(name="Arial", size=9, color="888888")
    s.fill = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row += 2

    # Column header
    def hcell(r, c, val):
        cl = ws.cell(row=r, column=c, value=val)
        cl.fill, cl.font = HEADER_FILL, HEADER_FONT
        cl.alignment = Alignment(horizontal="center", vertical="center")
        cl.border = THIN_BORDER

    hcell(row, 1, "")
    for i, yr in enumerate(years, start=2):
        hcell(row, i, yr)
    hcell(row, num_cols, "Total")
    row += 1

    # Data row formats for readability 
    even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_fill  = PatternFill(start_color="F7F8F9", end_color="F7F8F9", fill_type="solid")
    data_idx = 0

    for g in report_groups:
        if "section" in g:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            sc = ws.cell(row=row, column=1, value=g["section"])
            sc.fill = SECT_FILL
            sc.font = SECT_FONT
            sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 20
            row += 1
            data_idx = 0
        else:
            key = g["key"]
            row_fill = even_fill if data_idx % 2 == 0 else odd_fill

            lc = ws.cell(row=row, column=1, value=g["label"])
            lc.font = DATA_FONT
            lc.fill = row_fill
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            lc.border = HAIR_BOTTOM

            row_total = 0
            for i, yr in enumerate(years, start=2):
                val = counts[key].get(yr, 0)
                row_total += val
                vc = ws.cell(row=row, column=i, value=val)
                vc.font = DATA_FONT
                vc.fill = row_fill
                vc.alignment = Alignment(horizontal="center", vertical="center")
                # Use decimal format for acreage, integer for everything else
                vc.number_format = "#,##0.0" if key == "fire_acres" else "#,##0"
                vc.border = HAIR_BOTTOM

            tc = ws.cell(row=row, column=num_cols, value=row_total)
            tc.font = TOTAL_FONT
            tc.fill = row_fill
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.number_format = "#,##0.0" if key == "fire_acres" else "#,##0"
            tc.border = HAIR_BOTTOM

            row += 1
            data_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 38
    for ci in range(2, num_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

#  Prompt config - This is what the user fills in. Note that password is being passed as a string to avoid any issues across IDEs
def prompt_config():
    print("\n" + "=" * 60)
    print("  Combined AFG Grant Report — NFIRS + NERIS")
    print("=" * 60)

    print("\n── NERIS Credentials ──")
    neris_user = input("  NERIS Email: ").strip()
    print("  NERIS Password (characters will be visible):")
    neris_pass = input("  > ").strip()
    entity_id  = input("  NERIS Entity ID (e.g. FD26163151): ").strip()

    print("\n── NFIRS ArcGIS Filters ──")
    state = input("  State abbreviation (e.g. VA): ").strip().upper()
    fdid  = input("  FDID: ").strip()
    nfirs_years = [2022, 2023, 2024]

    if not all([neris_user, neris_pass, entity_id, state, fdid]):
        sys.exit("✗ All fields are required.")

    return {
        "neris_user": neris_user, "neris_pass": neris_pass,
        "entity_id": entity_id, "state": state, "fdid": fdid,
        "nfirs_years": nfirs_years,
    }


def main():
    ensure_dependencies()
    from openpyxl import Workbook

    cfg = prompt_config()

    # NFIRS
    print(f"\nFetching NFIRS — {cfg['state']} / FDID {cfg['fdid']} …")
    nfirs_records = fetch_nfirs_records(cfg["state"], cfg["fdid"], cfg["nfirs_years"])
    nfirs_years, nfirs_counts = build_nfirs_counts(nfirs_records)
    print(f"  NFIRS years: {', '.join(map(str, nfirs_years)) if nfirs_years else 'none'}")

    # Pull department name from NFIRS FD_NAME field
    nfirs_dept_name = ""
    for rec in nfirs_records:
        fd = (rec.get("FD_NAME") or "").strip()
        if fd:
            nfirs_dept_name = fd
            break

    # NERIS
    neris_client = neris_authenticate(cfg["neris_user"], cfg["neris_pass"])
    neris_dept_name = ""
    try:
        ent = neris_client.get_entity(cfg["entity_id"])
        neris_dept_name = ent.get("name", "") if isinstance(ent, dict) else ""
    except Exception:
        pass

    print(f"\n  NFIRS dept name: {nfirs_dept_name or '(none found)'}")
    print(f"  NERIS dept name: {neris_dept_name or '(none found)'}")

    neris_incidents = fetch_neris_incidents(neris_client, cfg["entity_id"])
    neris_years, neris_counts = build_neris_counts(neris_incidents)
    print(f"  NERIS years: {', '.join(map(str, neris_years)) if neris_years else 'none'}")

    # Create the workbook!
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "NFIRS"
    if nfirs_years:
        write_sheet(ws1, nfirs_years, nfirs_counts,
                    "NFIRS", f"{cfg['state']} / FDID {cfg['fdid']}",
                    nfirs_dept_name or cfg["fdid"],
                    NFIRS_REPORT_GROUPS)
    else:
        ws1.cell(row=1, column=1, value="No NFIRS records found for the selected filters.")

    ws2 = wb.create_sheet("NERIS")
    if neris_years:
        write_sheet(ws2, neris_years, neris_counts,
                    "NERIS", cfg["entity_id"],
                    neris_dept_name or cfg["entity_id"],
                    NERIS_REPORT_GROUPS)
    else:
        ws2.cell(row=1, column=1, value="No NERIS incidents found for this entity.")

    date_str = datetime.now().strftime("%Y-%m-%d")
    safe = "".join(c if c.isalnum() or c in " -_" else "" for c in (neris_dept_name or nfirs_dept_name or cfg["entity_id"])).strip()
    filename = os.path.join(
        os.path.expanduser("~"), "Documents",
        f"AFG_Combined_{safe}_{date_str}.xlsx",
    )
    if not os.path.isdir(os.path.dirname(filename)):
        filename = os.path.join(os.getcwd(), os.path.basename(filename))

    wb.save(filename)
    print(f"\n✓ Report saved: {filename}")
    print("\n" + "=" * 60)
    print("  PROCESS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
